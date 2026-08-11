from __future__ import annotations

import asyncio
import contextlib
import html
import io
import json
from html.parser import HTMLParser
import math
import random
import re
import sys
import time
from collections import deque
from pathlib import Path
from typing import Any
from urllib.parse import parse_qsl, urlencode, urljoin, urlparse, urlunparse

import aiohttp
from curl_cffi.requests import AsyncSession as CurlAsyncSession
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# НАСТРОЙКИ
# ============================================================

BASE_URL = "https://www.bauhaus.ee"
OUTPUT_FILE = Path("bauhaus_all_products.xlsx")
CATEGORY_CACHE_FILE = Path("bauhaus_categories.json")
EAN_CACHE_FILE = Path("bauhaus_ean_cache.json")
CATALOG_DATA_FILE = Path("bauhaus_catalog_cache.json")
STATISTICS_FILE = Path("bauhaus_statistics.json")

TEST_MODE = False

CATEGORY_CONCURRENCY = 1
PAGE_CONCURRENCY = 2
EAN_CONCURRENCY = 200

# Динамическая подстройка нагрузки.
EAN_FALLBACK_CONCURRENCY = 180
EAN_RECOVERY_STEP = 10
EAN_RECOVERY_INTERVAL = 10.0
EAN_MAX_CONCURRENCY = 200
EAN_WORKER_COUNT = EAN_MAX_CONCURRENCY
EAN_MAX_CLIENTS = EAN_MAX_CONCURRENCY
VERTICAL_LOG_INTERVAL = 2.0

# Последовательная обработка: сначала страницы одной категории, затем все её EAN.
# Следующая категория начинается только после полного опустошения EAN-очереди.

# Защита при сильном падении скорости EAN.
# Если скорость держится ниже порога, новые EAN-запросы останавливаются на паузу.
EAN_SPEED_PAUSE_THRESHOLD = 10.0
EAN_SPEED_PAUSE_SECONDS = 5.0
EAN_SPEED_LOW_CONFIRMATIONS = 3
EAN_SPEED_CHECK_INTERVAL = 1.0
EAN_SPEED_MIN_COMPLETED = 300
EAN_SPEED_WARMUP_SECONDS = 8.0
EAN_SPEED_MIN_ACTIVE_RATIO = 0.50

# После подтверждённого падения скорости лимит дополнительно уменьшается на шаг.
EAN_SPEED_REDUCE_STEP = 10

INVENTORY_CONCURRENCY = 5
INVENTORY_BATCH_SIZE = 40

MAX_RETRIES = 6
REQUEST_DELAY_MIN = 0.15
REQUEST_DELAY_MAX = 0.40
MAX_PAGES_PER_CATEGORY = 500
EMPTY_PAGE_RETRIES = 3
EMPTY_RETRY_DELAY_MIN = 3.0
EMPTY_RETRY_DELAY_MAX = 8.0

EAN_MAX_RETRIES = 4
EAN_REQUEST_TIMEOUT = 45
EAN_STREAM_CHUNK_SIZE = 32768
EAN_STREAM_MAX_BYTES = 800_000
EAN_SAVE_EVERY = 500

STORE_IDS = ("101", "103")
WEB_STORE_IDS = ("121", "105")

DASHBOARD_ACTIVE = False
HTTP_RETRIES = 0
HTTP_FINAL_ERRORS = 0
EMPTY_CATEGORY_PAGES = 0



HEADERS = {
    "accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "image/avif,image/webp,image/apng,*/*;q=0.8"
    ),
    "accept-language": "et-EE,et;q=0.9,ru-RU;q=0.8,ru;q=0.7,en;q=0.6",
    "cache-control": "no-cache",
    "pragma": "no-cache",
    "referer": BASE_URL + "/",
    "upgrade-insecure-requests": "1",
    "user-agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/150.0.0.0 Safari/537.36"
    ),
}

TOP_LEVEL_EXCLUDED_SLUGS = {
    "", "api", "artiklid-ja-napunaiteid", "blog", "brand", "brands",
    "ettevottest", "info", "kampaaniad", "kaubamajad", "kaubamargid",
    "kinkekaart", "kliendileht", "klienditugi", "kontakt", "login",
    "media", "oiguslik", "otsing", "profimuuk", "secure", "teenused",
    "tooriistade-laenutus",
}

BARCODE_KEYS = (
    "gtin", "gtin8", "gtin12", "gtin13", "gtin14",
    "ean", "ean8", "ean13", "ean14", "barcode",
)

# Только резервный поиск в данных Next.js, если JSON-LD отсутствует.
# Основной путь теперь: <script type="application/ld+json"> -> json.loads().
NEXT_DATA_BARCODE_PATTERN = re.compile(
    r'(?i)(?:\\?")(?:gtin(?:8|12|13|14)?|ean(?:8|13|14)?|barcode)(?:\\?")'
    r'\s*:\s*(?:\\?")(\d{8}|\d{12,14})(?:\\?")'
)


class AdjustableLimiter:
    """Лимитер, число одновременных задач которого меняется на лету."""

    def __init__(self, limit: int) -> None:
        self._limit = max(1, int(limit))
        self._active = 0
        self._condition = asyncio.Condition()

    @property
    def limit(self) -> int:
        return self._limit

    @property
    def active(self) -> int:
        return self._active

    async def set_limit(self, value: int) -> None:
        async with self._condition:
            self._limit = max(1, int(value))
            self._condition.notify_all()

    async def __aenter__(self) -> "AdjustableLimiter":
        async with self._condition:
            await self._condition.wait_for(
                lambda: self._active < self._limit
            )
            self._active += 1
        return self

    async def __aexit__(self, exc_type, exc, tb) -> None:
        async with self._condition:
            self._active = max(0, self._active - 1)
            self._condition.notify_all()


class AdaptiveLoadController:
    """Стартует с максимального EAN-limit и при проблемах только снижает его."""

    def __init__(
            self,
            ean_queue: asyncio.Queue,
            page_limiter: AdjustableLimiter,
            ean_limiter: AdjustableLimiter,
    ) -> None:
        self.ean_queue = ean_queue
        self.page_limiter = page_limiter
        self.ean_limiter = ean_limiter

        self.catalog_allowed = asyncio.Event()
        self.catalog_allowed.set()
        self.ean_allowed = asyncio.Event()
        self.ean_allowed.set()

        self.draining = False
        self.speed_paused = False
        self.speed_pauses = 0
        self.low_speed_checks = 0
        self.current_speed = 0.0
        self.completed_ean = 0
        self.category_ean_started_at = 0.0

        self.locked_limit: int | None = None
        self.last_raise_at = time.monotonic()
        self.restrictions = 0
        self.adaptations = 0
        self.last_reason = ""

        self._adaptation_lock = asyncio.Lock()
        self._pause_lock = asyncio.Lock()
        self._drain_task: asyncio.Task[None] | None = None
        self._speed_pause_task: asyncio.Task[None] | None = None

    async def wait_catalog_allowed(self) -> None:
        await self.catalog_allowed.wait()

    async def wait_ean_allowed(self) -> None:
        await self.ean_allowed.wait()

    async def begin_category_pages(self) -> None:
        """Перед новой категорией возвращает EAN-limit к 200 и блокирует EAN."""
        while self.speed_paused or self.draining:
            await asyncio.sleep(0.1)
        await self.ean_limiter.set_limit(EAN_CONCURRENCY)
        self.locked_limit = None
        self.low_speed_checks = 0
        self.current_speed = 0.0
        self.ean_allowed.clear()
        self.catalog_allowed.set()

    async def begin_category_ean(self) -> None:
        """После загрузки всех страниц разрешает только EAN текущей категории."""
        self.catalog_allowed.clear()
        self.ean_allowed.set()
        self.low_speed_checks = 0
        self.current_speed = 0.0
        self.completed_ean = 0
        self.category_ean_started_at = time.monotonic()

    async def finish_category(self) -> None:
        """Разрешает переход к страницам следующей категории."""
        while self.speed_paused or self.draining:
            await asyncio.sleep(0.1)
        self.ean_allowed.clear()
        self.catalog_allowed.set()
        self.low_speed_checks = 0
        self.current_speed = 0.0
        self.completed_ean = 0
        self.category_ean_started_at = 0.0

    def update_speed(self, speed: float, completed: int) -> None:
        self.current_speed = max(0.0, float(speed))
        self.completed_ean = max(0, int(completed))

    async def report_restriction(self, reason: str) -> None:
        """Вызывается только для HTTP 403/408/429 при EAN-запросе."""
        self.restrictions += 1
        self.last_reason = reason
        if self._drain_task is None or self._drain_task.done():
            self._drain_task = asyncio.create_task(
                self._adapt_after_restriction(reason)
            )

    async def _adapt_after_restriction(self, reason: str) -> None:
        async with self._adaptation_lock:
            if self.draining:
                return

            current = self.ean_limiter.limit
            target = max(EAN_FALLBACK_CONCURRENCY, current - EAN_RECOVERY_STEP)
            self.locked_limit = target
            self.draining = True
            self.adaptations += 1
            self.catalog_allowed.clear()
            self.ean_allowed.clear()
            await self.ean_limiter.set_limit(target)

            print(
                "\n" + "=" * 58 +
                f"\n[АВТОНАСТРОЙКА] {reason}" +
                f"\nEAN: {current} -> {target}" +
                "\nВсе новые запросы временно остановлены." +
                f"\nПауза {EAN_RECOVERY_INTERVAL:.0f} секунд." +
                "\n" + "=" * 58,
                flush=True,
            )

            await asyncio.sleep(EAN_RECOVERY_INTERVAL)
            self.draining = False
            self.last_raise_at = time.monotonic()
            self.ean_allowed.set()
            self.catalog_allowed.clear()

    async def _pause_all_for_low_speed(self) -> None:
        async with self._pause_lock:
            if self.speed_paused or self.draining:
                return

            self.speed_paused = True
            self.speed_pauses += 1
            old_limit = self.ean_limiter.limit
            new_limit = max(
                EAN_FALLBACK_CONCURRENCY,
                old_limit - EAN_SPEED_REDUCE_STEP,
            )
            await self.ean_limiter.set_limit(new_limit)
            self.catalog_allowed.clear()
            self.ean_allowed.clear()

            print(
                "\n" + "=" * 58 +
                "\n[ПАУЗА ПО СКОРОСТИ] EAN работает слишком медленно." +
                f"\nСкорость       : {self.current_speed:.1f} EAN/с" +
                f"\nПорог          : {EAN_SPEED_PAUSE_THRESHOLD:.1f} EAN/с" +
                f"\nEAN limit      : {old_limit} -> {new_limit}" +
                f"\nПолная пауза   : {EAN_SPEED_PAUSE_SECONDS:.0f} секунд" +
                "\nНовые страницы и новые EAN-запросы не запускаются." +
                "\n" + "=" * 58,
                flush=True,
            )

            await asyncio.sleep(EAN_SPEED_PAUSE_SECONDS)
            self.low_speed_checks = 0
            self.last_raise_at = time.monotonic()
            self.speed_paused = False
            self.ean_allowed.set()
            # Во время обработки EAN каталог остаётся заблокированным.
            self.catalog_allowed.clear()

            print(
                "\n" + "-" * 58 +
                f"\n[ПАУЗА ПО СКОРОСТИ] {EAN_SPEED_PAUSE_SECONDS:.0f} секунд завершены." +
                f"\nРабота продолжена с EAN-limit={new_limit}." +
                "\n" + "-" * 58,
                flush=True,
            )

    async def monitor(self) -> None:
        """Следит за скоростью: лимит никогда не повышается, только снижается."""
        while True:
            await asyncio.sleep(EAN_SPEED_CHECK_INTERVAL)
            has_work = self.ean_queue.qsize() > 0 or self.ean_limiter.active > 0
            enough_active = self.ean_limiter.active >= max(10, int(
                self.ean_limiter.limit * EAN_SPEED_MIN_ACTIVE_RATIO
            ))

            warmup_finished = (
                self.category_ean_started_at > 0
                and time.monotonic() - self.category_ean_started_at >= EAN_SPEED_WARMUP_SECONDS
            )

            speed_is_low = (
                has_work
                and enough_active
                and warmup_finished
                and self.completed_ean >= EAN_SPEED_MIN_COMPLETED
                and self.current_speed > 0
                and self.current_speed < EAN_SPEED_PAUSE_THRESHOLD
            )

            if speed_is_low and not self.speed_paused and not self.draining:
                self.low_speed_checks += 1
            else:
                self.low_speed_checks = 0

            if self.low_speed_checks >= EAN_SPEED_LOW_CONFIRMATIONS:
                if self._speed_pause_task is None or self._speed_pause_task.done():
                    self._speed_pause_task = asyncio.create_task(
                        self._pause_all_for_low_speed()
                    )


# ============================================================
# ФУНКЦИИ
# ============================================================

def clean_text(value: Any) -> str:
    if value is None:
        return ""

    text = html.unescape(str(value))
    return re.sub(r"\s+", " ", text).strip()


def clean_sku(value: Any) -> str:
    if isinstance(value, (list, tuple)):
        for item in value:
            sku = clean_text(item)
            if sku:
                return sku
        return ""
    return clean_text(value)


def as_number(value: Any) -> float | None:
    if value is None or value == "":
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = (
        str(value)
        .strip()
        .replace("\xa0", "")
        .replace("€", "")
        .replace(",", ".")
    )

    try:
        return float(text)
    except ValueError:
        return None


def nested_get(data: Any, *keys: str, default: Any = None) -> Any:
    current = data

    for key in keys:
        if not isinstance(current, dict):
            return default

        current = current.get(key)

        if current is None:
            return default

    return current


def normalize_url(url: str) -> str:
    if not url:
        return ""

    absolute = urljoin(BASE_URL, html.unescape(url.strip()))
    parsed = urlparse(absolute)

    if parsed.netloc not in {"bauhaus.ee", "www.bauhaus.ee"}:
        return ""

    path = re.sub(r"/+", "/", parsed.path).rstrip("/") or "/"

    return urlunparse(
        (
            "https",
            "www.bauhaus.ee",
            path,
            "",
            "",
            "",
        )
    )


def add_query_parameter(url: str, name: str, value: Any) -> str:
    parsed = urlparse(url)
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query[name] = str(value)

    return urlunparse(
        (
            parsed.scheme,
            parsed.netloc,
            parsed.path,
            parsed.params,
            urlencode(query),
            parsed.fragment,
        )
    )


def chunked(items: list[str], size: int) -> list[list[str]]:
    return [
        items[index:index + size]
        for index in range(0, len(items), size)
    ]


async def request_text(
        session: aiohttp.ClientSession,
        url: str,
        semaphore: asyncio.Semaphore,
        *,
        params: list[tuple[str, str]] | dict[str, str] | None = None,
        headers: dict[str, str] | None = None,
) -> str | None:
    global HTTP_RETRIES, HTTP_FINAL_ERRORS

    async with semaphore:
        for attempt in range(1, MAX_RETRIES + 1):
            await asyncio.sleep(
                random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX)
            )

            try:
                async with session.get(
                        url,
                        params=params,
                        headers=headers,
                        allow_redirects=True,
                        timeout=aiohttp.ClientTimeout(total=90),
                ) as response:
                    if response.status == 200:
                        return await response.text(errors="replace")

                    if response.status in {403, 408, 429, 500, 502, 503, 504}:
                        HTTP_RETRIES += 1
                        wait_time = min(
                            60,
                            2 ** attempt + random.uniform(0.5, 3.0),
                        )

                        if not DASHBOARD_ACTIVE:
                            print(
                                f"  HTTP {response.status}: {url}\n"
                                f"  Повтор {attempt}/{MAX_RETRIES}, "
                                f"ждём {wait_time:.1f} сек."
                            )

                        await asyncio.sleep(wait_time)
                        continue

                    HTTP_FINAL_ERRORS += 1
                    if not DASHBOARD_ACTIVE:
                        print(f"  HTTP {response.status}: {url}")
                    return None

            except (aiohttp.ClientError, asyncio.TimeoutError) as error:
                HTTP_RETRIES += 1
                wait_time = min(
                    60,
                    2 ** attempt + random.uniform(0.5, 3.0),
                )

                if not DASHBOARD_ACTIVE:
                    error_text = str(error).strip() or type(error).__name__
                    print(
                        f"  Ошибка запроса: {error_text}\n"
                        f"  {url}\n"
                        f"  Повтор {attempt}/{MAX_RETRIES}, "
                        f"ждём {wait_time:.1f} сек."
                    )

                await asyncio.sleep(wait_time)

    HTTP_FINAL_ERRORS += 1

    if not DASHBOARD_ACTIVE:
        print(f"  Не удалось загрузить: {url}")

    return None


def document_versions(text: str) -> list[str]:
    """
    Создаёт варианты документа:
    - исходный HTML;
    - HTML entities decoded;
    - строки Next.js, извлечённые из self.__next_f.push(...).
    """
    versions: list[str] = [text]

    decoded_html = html.unescape(text)

    if decoded_html != text:
        versions.append(decoded_html)

    for match in re.finditer(
            r'self\.__next_f\.push\(\[\d+,\s*("(?:\\.|[^"\\])*")',
            text,
            flags=re.DOTALL,
    ):
        encoded = match.group(1)

        try:
            decoded = json.loads(encoded)
        except json.JSONDecodeError:
            continue

        if isinstance(decoded, str):
            versions.append(decoded)

    expanded: list[str] = []

    for version in versions:
        expanded.append(version)
        expanded.append(
            version
            .replace('\\"', '"')
            .replace("\\/", "/")
            .replace("\\u0026", "&")
            .replace("\\u003c", "<")
            .replace("\\u003e", ">")
            .replace("\\u0022", '"')
        )

    unique: list[str] = []
    seen: set[str] = set()

    for version in expanded:
        if version and version not in seen:
            seen.add(version)
            unique.append(version)

    return unique


def json_values_after_marker(text: str, marker: str) -> list[Any]:
    decoder = json.JSONDecoder()
    values: list[Any] = []
    start = 0

    while True:
        position = text.find(marker, start)

        if position == -1:
            break

        value_start = position + len(marker)

        while (
                value_start < len(text)
                and text[value_start] in " \r\n\t"
        ):
            value_start += 1

        try:
            value, _ = decoder.raw_decode(text[value_start:])
            values.append(value)
        except json.JSONDecodeError:
            pass

        start = position + len(marker)

    return values


def is_product_hit(item: Any) -> bool:
    if not isinstance(item, dict):
        return False

    sku = clean_sku(item.get("sku"))

    if not sku:
        return False

    return any(
        key in item
        for key in (
            "name",
            "url",
            "canonical_url",
            "image_url",
            "grid_image",
            "price",
            "bauhaus_price",
        )
    )


def extract_hits_from_document(text: str) -> list[dict[str, Any]]:
    found: dict[str, dict[str, Any]] = {}

    for version in document_versions(text):
        for value in json_values_after_marker(version, '"hits":'):
            if not isinstance(value, list):
                continue

            for item in value:
                if not is_product_hit(item):
                    continue

                sku = clean_sku(item.get("sku"))

                if sku:
                    found[sku] = item

    return list(found.values())


def extract_integer_values(text: str, field_name: str) -> list[int]:
    values: list[int] = []

    for version in document_versions(text):
        for match in re.finditer(
                rf'"{re.escape(field_name)}"\s*:\s*(\d+)',
                version,
        ):
            try:
                values.append(int(match.group(1)))
            except ValueError:
                pass

    return values


def extract_catalog_metadata(text: str) -> dict[str, int]:
    return {
        "nb_pages": max(
            extract_integer_values(text, "nbPages"),
            default=0,
        ),
        "nb_hits": max(
            extract_integer_values(text, "nbHits"),
            default=0,
        ),
        "hits_per_page": max(
            extract_integer_values(text, "hitsPerPage"),
            default=0,
        ),
    }


def is_catalog_category_node(value: Any) -> bool:
    """Проверяет элемент дерева категорий BAUHAUS."""
    if not isinstance(value, dict):
        return False

    url_path = clean_text(value.get("url_path")).strip("/")
    children = value.get("children")

    if not url_path or "/" in url_path:
        return False

    if not isinstance(children, list) or not children:
        return False

    return any(
        isinstance(child, dict)
        and clean_text(child.get("url_path")).startswith(url_path + "/")
        for child in children
    )


def extract_category_tree(document: str) -> list[dict[str, Any]]:
    """
    Извлекает настоящее дерево меню из Next.js/RSC.

    В RSC BAUHAUS дерево передаётся как:
        "categories": [{"name": ..., "url_path": ..., "children": [...]}, ...]

    В документе могут присутствовать и другие поля categories — например,
    категории карточек товаров. Поэтому выбирается самый большой массив,
    состоящий из корневых узлов с url_path и дочерними категориями.
    """
    best: list[dict[str, Any]] = []

    for version in document_versions(document):
        for value in json_values_after_marker(version, '"categories":'):
            if not isinstance(value, list):
                continue

            roots = [item for item in value if is_catalog_category_node(item)]

            if len(roots) > len(best):
                best = roots

    return best


def extract_deployment_id(document: str) -> str:
    """Берёт текущий Next.js deployment id из HTML, если он присутствует."""
    matches = re.findall(r"dpl_[A-Za-z0-9_-]+", document)
    return matches[0] if matches else ""


async def request_home_rsc(
        session: aiohttp.ClientSession,
        semaphore: asyncio.Semaphore,
        home_document: str,
        attempt_number: int,
) -> str | None:
    """Запрашивает общий RSC-сегмент, содержащий меню категорий."""
    token = "".join(
        random.choice("abcdefghijklmnopqrstuvwxyz0123456789")
        for _ in range(5)
    )

    rsc_headers = {
        **HEADERS,
        "accept": "*/*",
        "rsc": "1",
        "next-url": "/",
        "next-router-prefetch": "1",
        "next-router-segment-prefetch": "/!KGNvbW1vbik",
        "referer": BASE_URL + "/",
    }

    deployment_id = extract_deployment_id(home_document)
    if deployment_id:
        rsc_headers["x-deployment-id"] = deployment_id

    # Параметр _rsc специально делается новым при каждой попытке,
    # чтобы CDN не возвращал один и тот же неполный сегмент.
    return await request_text(
        session,
        BASE_URL + "/",
        semaphore,
        params={"_rsc": token},
        headers=rsc_headers,
    )


def category_urls_from_tree(
        tree: list[dict[str, Any]],
) -> tuple[list[tuple[str, str, int]], list[str], dict[str, Any]]:
    """
    Рекурсивно обходит дерево категорий на ЛЮБУЮ глубину.

    Возвращает:
      1. Корневые разделы для вывода в консоль.
      2. URL всех листовых категорий для выгрузки.
      3. Аудит дерева: глубина, все узлы и связи родитель -> ребёнок.

    Листовой считается категория, у которой нет корректных дочерних
    категорий. Именно листовые категории скачиваются, чтобы крупные разделы
    не упирались в ограничение Algolia примерно в 5000 товаров.
    """
    roots: list[tuple[str, str, int]] = []
    download_urls: list[str] = []
    seen_roots: set[str] = set()
    seen_downloads: set[str] = set()
    seen_nodes: set[str] = set()
    audit_nodes: list[dict[str, Any]] = []
    depth_counts: dict[int, int] = {}
    max_depth = 0

    def valid_node(node: Any) -> bool:
        return (
                isinstance(node, dict)
                and bool(clean_text(node.get("url_path")).strip("/"))
        )

    def walk(
            node: dict[str, Any],
            *,
            depth: int,
            parent_url: str = "",
            parent_path: str = "",
    ) -> None:
        nonlocal max_depth

        url_path = clean_text(node.get("url_path")).strip("/")
        if not url_path:
            return

        url = normalize_url("/" + url_path)
        if not url:
            return

        # Защита от циклов или повторно вложенного одного и того же узла.
        if url in seen_nodes:
            return
        seen_nodes.add(url)

        name = clean_text(node.get("name")) or url_path.rsplit("/", 1)[-1]
        children_raw = node.get("children")
        valid_children = (
            [child for child in children_raw if valid_node(child)]
            if isinstance(children_raw, list)
            else []
        )

        current_path = f"{parent_path} > {name}" if parent_path else name
        is_leaf = not valid_children

        max_depth = max(max_depth, depth)
        depth_counts[depth] = depth_counts.get(depth, 0) + 1
        audit_nodes.append(
            {
                "name": name,
                "url": url,
                "url_path": url_path,
                "depth": depth,
                "parent_url": parent_url,
                "category_path": current_path,
                "children_count": len(valid_children),
                "is_leaf": is_leaf,
            }
        )

        if is_leaf:
            if url not in seen_downloads:
                seen_downloads.add(url)
                download_urls.append(url)
            return

        # Никакого ограничения по количеству уровней здесь нет.
        for child in valid_children:
            walk(
                child,
                depth=depth + 1,
                parent_url=url,
                parent_path=current_path,
            )

    for node in tree:
        if not valid_node(node):
            continue

        url_path = clean_text(node.get("url_path")).strip("/")
        if "/" in url_path:
            continue
        if url_path.lower() in TOP_LEVEL_EXCLUDED_SLUGS:
            continue

        url = normalize_url("/" + url_path)
        if not url or url in seen_roots:
            continue

        children = node.get("children")
        valid_children = (
            [child for child in children if valid_node(child)]
            if isinstance(children, list)
            else []
        )
        if not valid_children:
            continue

        seen_roots.add(url)
        name = clean_text(node.get("name")) or url_path
        roots.append((url, name, len(valid_children)))
        walk(node, depth=0)

    audit = {
        "root_categories": len(roots),
        "all_category_nodes": len(audit_nodes),
        "leaf_categories": len(download_urls),
        "max_depth": max_depth,
        "nodes_by_depth": {
            str(depth): count
            for depth, count in sorted(depth_counts.items())
        },
        "nodes": audit_nodes,
    }

    return roots, download_urls, audit


async def discover_top_level_categories(
        session: aiohttp.ClientSession,
        page_semaphore: asyncio.Semaphore,
        home_document: str,
) -> list[str]:
    """
    Получает главные разделы напрямую из дерева меню Next.js/RSC.

    Никакие произвольные строки HTML больше не превращаются в URL, поэтому
    запросы вида `/><script`, `/><meta`, `$` и ссылки товаров исключены.
    """
    tree = extract_category_tree(home_document)

    # В обычном HTML дерево бывает не всегда. Запрашиваем несколько RSC-
    # сегментов и останавливаемся сразу после нахождения полного меню.
    if not tree:
        print("\nВ HTML дерево категорий не найдено. Запрашиваем Next.js RSC...")

        for attempt in range(1, 6):
            rsc_document = await request_home_rsc(
                session,
                page_semaphore,
                home_document,
                attempt,
            )

            if not rsc_document:
                print(f"  RSC {attempt}/5: ответа нет")
                continue

            candidate_tree = extract_category_tree(rsc_document)
            print(
                f"  RSC {attempt}/5: "
                f"корневых категорий найдено {len(candidate_tree)}"
            )

            if len(candidate_tree) > len(tree):
                tree = candidate_tree

            # На момент исследования основное меню содержало больше десяти
            # корневых разделов. Порог не задаёт сами URL и лишь позволяет
            # не делать лишние запросы после получения полного дерева.
            if len(tree) >= 10:
                break

    rows, categories, category_audit = category_urls_from_tree(tree)

    if rows:
        print(f"\nГлавных товарных разделов из RSC: {len(rows)}")
        for url, name, child_count in rows:
            print(f"  ✓ {name}: {url} | подразделов: {child_count}")
        print(
            f"\nДля полной выгрузки выбрано листовых подразделов: "
            f"{len(categories)}"
        )
        print(
            "Глубина дерева категорий: "
            f"{category_audit.get('max_depth', 0)} | "
            "всего узлов: "
            f"{category_audit.get('all_category_nodes', 0)}"
        )
        print(
            "Узлов по уровням: "
            f"{category_audit.get('nodes_by_depth', {})}"
        )

    # Кэш нужен только при временной недоступности RSC. Новый неполный
    # результат не должен затирать ранее сохранённый полный список.
    if not categories and CATEGORY_CACHE_FILE.exists():
        try:
            cached = json.loads(CATEGORY_CACHE_FILE.read_text(encoding="utf-8"))
            if isinstance(cached, list):
                categories = [
                    normalize_url(str(url))
                    for url in cached
                    if normalize_url(str(url))
                ]
                categories = list(dict.fromkeys(categories))
                if categories:
                    print(
                        "\nВнимание: RSC-меню временно не получено. "
                        "Используем сохранённый список разделов."
                    )
        except (OSError, json.JSONDecodeError):
            pass

    if TEST_MODE:
        categories = categories[:1]

    print(f"\nНайдено категорий для выгрузки: {len(categories)}")
    return categories


def extract_category_data(
        hit: dict[str, Any],
) -> tuple[str, str, str, str]:
    categories = hit.get("categories")

    level0 = ""
    level1 = ""
    level2 = ""
    path = ""

    if isinstance(categories, dict):
        def first(key: str) -> str:
            value = categories.get(key)

            if isinstance(value, list) and value:
                return clean_text(value[0])

            if isinstance(value, str):
                return clean_text(value)

            return ""

        raw0 = first("level0")
        raw1 = first("level1")
        raw2 = first("level2")

        level0 = raw0.split(" /// ")[-1] if raw0 else ""
        level1 = raw1.split(" /// ")[-1] if raw1 else ""
        level2 = raw2.split(" /// ")[-1] if raw2 else ""

        deepest = raw2 or raw1 or raw0
        path = deepest.replace(" /// ", " > ")

    elif isinstance(categories, list):
        names: list[str] = []

        for category in categories:
            if not isinstance(category, dict):
                continue

            name = clean_text(category.get("name"))

            if name and name.lower() != "kampaaniad":
                names.append(name)

        names = list(dict.fromkeys(names))

        if names:
            level0 = names[0]

        if len(names) > 1:
            level1 = names[1]

        if len(names) > 2:
            level2 = names[2]

        path = " > ".join(names)

    return level0, level1, level2, path


def extract_price_data(hit: dict[str, Any]) -> dict[str, Any]:
    result = {
        "price": None,
        "ordinary_price": None,
        "regular_price": None,
        "discount": None,
        "unit_price": None,
        "unit": "",
        "currency": "EUR",
        "campaign_date_to": "",
        "merchandising_type": "",
        "merchandising_badge_type": "",
    }

    bauhaus_price = hit.get("bauhaus_price")

    if isinstance(bauhaus_price, dict):
        result["price"] = as_number(
            nested_get(bauhaus_price, "final_price", "value")
        )
        result["ordinary_price"] = as_number(
            nested_get(bauhaus_price, "ordinary_price", "value")
        )
        result["regular_price"] = as_number(
            nested_get(bauhaus_price, "regular_price", "value")
        )
        result["discount"] = as_number(
            nested_get(bauhaus_price, "discount", "amount_off")
        )
        result["unit_price"] = as_number(
            nested_get(bauhaus_price, "unit_price", "value")
        )
        result["unit"] = clean_text(
            nested_get(bauhaus_price, "unit_price", "unit", default="")
        )
        result["currency"] = clean_text(
            nested_get(
                bauhaus_price,
                "final_price",
                "currency",
                default="EUR",
            )
        ) or "EUR"
        result["campaign_date_to"] = clean_text(
            bauhaus_price.get("campaign_date_to")
        )
        result["merchandising_type"] = clean_text(
            bauhaus_price.get("merchandising_type")
        )
        result["merchandising_badge_type"] = clean_text(
            bauhaus_price.get("merchandising_badge_type")
        )

        return result

    price = hit.get("price")

    if isinstance(price, dict):
        eur = price.get("EUR")

        if isinstance(eur, dict):
            result["price"] = as_number(
                eur.get("group_0", eur.get("default"))
            )

    result["campaign_date_to"] = clean_text(
        hit.get("campaign_price_to_date")
    )
    result["merchandising_type"] = clean_text(
        hit.get("merchandising_type")
    )
    result["merchandising_badge_type"] = clean_text(
        hit.get("merchandising_badge_type")
    )

    return result


def product_from_hit(
        hit: dict[str, Any],
        source_category_url: str,
) -> dict[str, Any] | None:
    sku = clean_sku(hit.get("sku"))

    if not sku:
        return None

    product_url = clean_text(
        hit.get("url")
        or hit.get("canonical_url")
        or hit.get("url_key")
    )

    if product_url:
        product_url = urljoin(
            BASE_URL + "/",
            product_url.lstrip("/"),
        )

    image_url = clean_text(
        hit.get("image_url")
        or hit.get("thumbnail_url")
        or nested_get(hit, "grid_image", "url", default="")
    )

    retina_image_url = clean_text(
        hit.get("double_density_image_url")
        or nested_get(hit, "grid_image", "url_retina", default="")
    )

    level0, level1, level2, category_path = extract_category_data(hit)
    price = extract_price_data(hit)

    rating = None
    review_count = None
    rating_summary = hit.get("rating_summary")

    if isinstance(rating_summary, dict):
        rating = as_number(rating_summary.get("average_score"))
        review_count = rating_summary.get("review_count")
    else:
        rating = as_number(rating_summary)

    return {
        "sku": sku,
        "ean": clean_text(hit.get("ean")),
        "name": clean_text(hit.get("name")),
        "brand": clean_text(
            hit.get("brand_name")
            or hit.get("brand")
        ),
        "price": price["price"],
        "ordinary_price": price["ordinary_price"],
        "regular_price": price["regular_price"],
        "discount": price["discount"],
        "unit_price": price["unit_price"],
        "unit": price["unit"],
        "currency": price["currency"],
        "stock_status": clean_text(hit.get("stock_status")),
        "availability": hit.get("availability"),
        "rating": rating,
        "review_count": review_count,
        "category_level_0": level0,
        "category_level_1": level1,
        "category_level_2": level2,
        "category_path": category_path,
        "campaign_date_to": price["campaign_date_to"],
        "merchandising_type": price["merchandising_type"],
        "merchandising_badge_type": price[
            "merchandising_badge_type"
        ],
        "image_url": image_url,
        "retina_image_url": retina_image_url,
        "product_url": product_url,
        "source_category_url": source_category_url,
        "object_id": clean_text(
            hit.get("objectID")
            or hit.get("id")
        ),
    }


async def fetch_category_page(
        session: aiohttp.ClientSession,
        page_semaphore: asyncio.Semaphore,
        category_url: str,
        page_number: int,
        *,
        retry_if_empty: bool = False,
) -> tuple[int, list[dict[str, Any]], dict[str, int]]:
    """Загружает страницу категории.

    Если страница должна существовать по серверному nbPages, но неожиданно
    пришла пустой, делает несколько повторов с увеличенной паузой. Это
    помогает при временной защите/нестабильном ответе сайта.
    """
    url = category_url

    if page_number > 1:
        url = add_query_parameter(url, "page", page_number)

    attempts = EMPTY_PAGE_RETRIES + 1 if retry_if_empty else 1
    last_metadata = {
        "nb_pages": 0,
        "nb_hits": 0,
        "hits_per_page": 0,
    }

    for attempt in range(1, attempts + 1):
        document = await request_text(session, url, page_semaphore)

        if document:
            hits = extract_hits_from_document(document)
            last_metadata = extract_catalog_metadata(document)

            if hits:
                return page_number, hits, last_metadata

        if attempt < attempts:
            delay = random.uniform(
                EMPTY_RETRY_DELAY_MIN * attempt,
                EMPTY_RETRY_DELAY_MAX * attempt,
            )
            print(
                f"  Страница {page_number} неожиданно пустая. "
                f"Повтор {attempt}/{EMPTY_PAGE_RETRIES} через {delay:.1f} сек."
            )
            await asyncio.sleep(delay)

    return page_number, [], last_metadata


class JsonLdScriptParser(HTMLParser):
    """Извлекает только JSON-LD, не разбирая всё DOM-дерево страницы."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=False)
        self._inside_jsonld = False
        self._buffer: list[str] = []
        self.blocks: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() != "script":
            return
        attributes = {str(key).lower(): (value or "") for key, value in attrs}
        script_type = attributes.get("type", "").lower().split(";", 1)[0].strip()
        if script_type == "application/ld+json":
            self._inside_jsonld = True
            self._buffer = []

    def handle_data(self, data: str) -> None:
        if self._inside_jsonld:
            self._buffer.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "script" and self._inside_jsonld:
            block = "".join(self._buffer).strip()
            if block:
                self.blocks.append(block)
            self._inside_jsonld = False
            self._buffer = []


def walk_json(value: Any):
    """Итеративный обход JSON без риска упереться в глубину рекурсии."""
    stack = [value]
    while stack:
        current = stack.pop()
        yield current
        if isinstance(current, dict):
            stack.extend(current.values())
        elif isinstance(current, list):
            stack.extend(current)


def json_type_contains_product(value: Any) -> bool:
    if isinstance(value, str):
        return value.lower() == "product"
    if isinstance(value, list):
        return any(isinstance(item, str) and item.lower() == "product" for item in value)
    return False


def barcode_from_mapping(mapping: dict[str, Any]) -> str:
    lowered = {str(key).lower(): value for key, value in mapping.items()}
    for key in BARCODE_KEYS:
        candidate = normalize_barcode_candidate(lowered.get(key))
        if candidate:
            return candidate
    return ""


def extract_ean_from_jsonld(page_html: str, expected_sku: str = "") -> str:
    """
    Находит Product JSON-LD и возвращает проверенный GTIN/EAN.

    Сначала предпочитает Product с совпадающим SKU. Если SKU в JSON-LD
    отсутствует, использует первый валидный Product GTIN.
    """
    parser = JsonLdScriptParser()
    try:
        parser.feed(page_html)
        parser.close()
    except Exception:
        return ""

    expected_sku = clean_text(expected_sku)
    fallback = ""

    for raw_block in parser.blocks:
        raw_block = html.unescape(raw_block).strip()
        try:
            payload = json.loads(raw_block)
        except json.JSONDecodeError:
            continue

        for node in walk_json(payload):
            if not isinstance(node, dict):
                continue
            if not json_type_contains_product(node.get("@type")):
                continue

            ean = barcode_from_mapping(node)
            if not ean:
                continue

            node_sku = clean_text(node.get("sku"))
            if expected_sku and node_sku and node_sku == expected_sku:
                return ean
            if not fallback:
                fallback = ean

    return fallback


def extract_ean_from_next_data(page_html: str) -> str:
    """Резерв: ищет gtin/ean в Next.js-данных, если JSON-LD не сработал."""
    match = NEXT_DATA_BARCODE_PATTERN.search(page_html)
    if not match:
        return ""
    return normalize_barcode_candidate(match.group(1))


def normalize_barcode_candidate(value: Any) -> str:
    """Оставляет только цифровой EAN/UPC/GTIN допустимой длины."""
    if value is None:
        return ""

    if isinstance(value, (int, float)):
        if isinstance(value, float) and not value.is_integer():
            return ""
        text = str(int(value))
    else:
        text = clean_text(value)

    digits = re.sub(r"\D", "", text)

    if len(digits) not in {8, 12, 13, 14}:
        return ""

    return digits if is_valid_gtin(digits) else ""


def is_valid_gtin(code: str) -> bool:
    """Проверяет контрольную цифру EAN-8, UPC-A, EAN-13 или GTIN-14."""
    if not code.isdigit() or len(code) not in {8, 12, 13, 14}:
        return False

    body = code[:-1]
    expected = int(code[-1])

    total = 0
    # Справа налево веса 3, 1, 3, 1...
    for index, char in enumerate(reversed(body)):
        weight = 3 if index % 2 == 0 else 1
        total += int(char) * weight

    calculated = (10 - total % 10) % 10
    return calculated == expected


def load_ean_cache() -> dict[str, dict[str, str]]:
    if not EAN_CACHE_FILE.exists():
        return {}

    try:
        data = json.loads(EAN_CACHE_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}

    if not isinstance(data, dict):
        return {}

    result: dict[str, dict[str, str]] = {}
    for sku, item in data.items():
        if isinstance(item, dict):
            result[str(sku)] = {
                "ean": clean_text(item.get("ean")),
                "source": clean_text(item.get("source")),
                "url": clean_text(item.get("url")),
            }

    return result


def save_ean_cache(cache: dict[str, dict[str, str]]) -> None:
    temporary = EAN_CACHE_FILE.with_suffix(".tmp")
    temporary.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporary.replace(EAN_CACHE_FILE)


async def fetch_product_ean(
        session: CurlAsyncSession,
        semaphore: AdjustableLimiter,
        controller: AdaptiveLoadController,
        sku: str,
        product_url: str,
) -> tuple[str, str, str, str]:
    """
    Быстро получает EAN из Product JSON-LD карточки.

    HTML всё ещё скачивается, потому что отдельного товарного API с gtin
    в исследованных запросах не найдено. Но вместо поиска таблиц и разбора
    всего документа извлекается только application/ld+json и выполняется
    обычный json.loads().
    """
    async with semaphore:
        for attempt in range(1, EAN_MAX_RETRIES + 1):
            try:
                # Очень маленький джиттер не даёт всем 200 задачам стартовать
                # в одну и ту же миллисекунду.
                await asyncio.sleep(random.uniform(0.005, 0.025))

                response = await session.get(
                    product_url,
                    headers={
                        **HEADERS,
                        "accept": "text/html,application/xhtml+xml",
                        "referer": BASE_URL + "/",
                    },
                    timeout=EAN_REQUEST_TIMEOUT,
                    allow_redirects=True,
                )

                resolved_url = str(response.url)

                if response.status_code == 200:
                    page_html = response.text

                    ean = extract_ean_from_jsonld(page_html, expected_sku=sku)
                    if ean:
                        return sku, ean, "jsonld_gtin", resolved_url

                    # Резерв для страниц, где Product JSON-LD временно
                    # отсутствует, но gtin остаётся в данных Next.js.
                    ean = extract_ean_from_next_data(page_html)
                    if ean:
                        return sku, ean, "next_data_gtin", resolved_url

                    return sku, "", "ean_not_found", resolved_url

                if response.status_code in {403, 408, 429, 500, 502, 503, 504}:
                    if response.status_code in {403, 408, 429}:
                        await controller.report_restriction(
                            f"HTTP {response.status_code} при EAN-запросе"
                        )

                    retry_after = response.headers.get("retry-after")
                    try:
                        wait = float(retry_after) if retry_after else 0.0
                    except (TypeError, ValueError):
                        wait = 0.0

                    wait = max(
                        wait,
                        min(30.0, 1.8 ** attempt + random.uniform(0.5, 1.8)),
                    )
                    await asyncio.sleep(wait)
                    continue

                return sku, "", f"http_{response.status_code}", resolved_url

            except Exception as error:
                if attempt == EAN_MAX_RETRIES:
                    return (
                        sku,
                        "",
                        f"request_failed:{type(error).__name__}",
                        product_url,
                    )
                await asyncio.sleep(
                    min(30.0, 1.8 ** attempt + random.uniform(0.5, 1.8))
                )

    return sku, "", "request_failed", product_url


async def fetch_inventory_batch(
        session: aiohttp.ClientSession,
        semaphore: asyncio.Semaphore,
        sku_batch: list[str],
        batch_number: int,
        batch_total: int,
) -> dict[str, dict[str, Any]]:
    params: list[tuple[str, str]] = [("type", "category")]

    for sku in sku_batch:
        params.append(("skus[]", str(sku)))

    for store_id in WEB_STORE_IDS:
        params.append(("web_s[]", store_id))

    for store_id in STORE_IDS:
        params.append(("str_s[]", store_id))

    inventory_headers = {
        **HEADERS,
        "accept": "*/*",
        "referer": BASE_URL + "/",
    }

    text = await request_text(
        session,
        BASE_URL + "/api/inventory/card",
        semaphore,
        params=params,
        headers=inventory_headers,
    )

    if not text:
        if not DASHBOARD_ACTIVE:
            print(
                f"  Остатки {batch_number}/{batch_total}: "
                "запрос не удался"
            )
        return {}

    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        if not DASHBOARD_ACTIVE:
            print(
                f"  Остатки {batch_number}/{batch_total}: "
                "ответ не JSON"
            )
        return {}

    response = data.get("response", {})
    if not isinstance(response, dict):
        print(
            f"  Остатки {batch_number}/{batch_total}: "
            f"поле response имеет тип {type(response).__name__}"
        )
        return {}

    raw_web_data = response.get("web", {})
    raw_physical_data = response.get("physic", {})

    # Иногда BAUHAUS возвращает web/physic не словарём по SKU,
    # а списком объектов. Приводим оба варианта к единому виду.
    web_data: dict[str, dict[str, Any]] = {}
    if isinstance(raw_web_data, dict):
        for key, value in raw_web_data.items():
            sku_key = str(key)
            if isinstance(value, dict):
                web_data[sku_key] = value
            elif isinstance(value, list) and value and isinstance(value[0], dict):
                web_data[sku_key] = value[0]
    elif isinstance(raw_web_data, list):
        for item in raw_web_data:
            if not isinstance(item, dict):
                continue
            sku_key = clean_text(
                item.get("sku")
                or item.get("product_sku")
                or item.get("article")
                or item.get("id")
            )
            if sku_key:
                web_data[sku_key] = item

    physical_data: dict[str, list[dict[str, Any]]] = {}
    if isinstance(raw_physical_data, dict):
        for key, value in raw_physical_data.items():
            sku_key = str(key)
            if isinstance(value, list):
                physical_data[sku_key] = [
                    item for item in value if isinstance(item, dict)
                ]
            elif isinstance(value, dict):
                physical_data[sku_key] = [value]
    elif isinstance(raw_physical_data, list):
        for item in raw_physical_data:
            if not isinstance(item, dict):
                continue
            sku_key = clean_text(
                item.get("sku")
                or item.get("product_sku")
                or item.get("article")
                or item.get("id")
            )
            if sku_key:
                physical_data.setdefault(sku_key, []).append(item)

    if isinstance(raw_web_data, list) or isinstance(raw_physical_data, list):
        print(
            f"  Остатки {batch_number}/{batch_total}: "
            "API вернул список вместо словаря — ответ нормализован"
        )

    result: dict[str, dict[str, Any]] = {}

    for raw_sku in sku_batch:
        sku = str(raw_sku)
        web_item = web_data.get(sku, {})
        physical_items = physical_data.get(sku, [])

        store_qty: dict[str, float | None] = {
            store_id: None
            for store_id in STORE_IDS
        }

        if isinstance(physical_items, list):
            for item in physical_items:
                if not isinstance(item, dict):
                    continue

                stock_id = clean_text(item.get("stock_id"))

                if stock_id in store_qty:
                    store_qty[stock_id] = as_number(
                        item.get("qty")
                    )

        known = [
            value
            for value in store_qty.values()
            if value is not None
        ]

        result[sku] = {
            "web_qty": as_number(web_item.get("qty")),
            "web_status": clean_text(web_item.get("status")),
            "stock_101": store_qty.get("101"),
            "stock_103": store_qty.get("103"),
            "physical_total": (
                sum(known)
                if known
                else None
            ),
            "inventory_received": (
                    sku in web_data
                    or sku in physical_data
            ),
        }

    if not DASHBOARD_ACTIVE:
        print(
            f"  Остатки {batch_number}/{batch_total}: "
            f"{len(sku_batch)} SKU"
        )

    return result


async def download_all_inventory(
        session: aiohttp.ClientSession,
        semaphore: asyncio.Semaphore,
        sku_list: list[str],
) -> dict[str, dict[str, Any]]:
    global DASHBOARD_ACTIVE
    DASHBOARD_ACTIVE = True

    batches = chunked(
        sku_list,
        INVENTORY_BATCH_SIZE,
    )

    total_skus = len(sku_list)
    completed_batches = 0
    completed_skus = 0
    inventory: dict[str, dict[str, Any]] = {}
    started_at = time.monotonic()
    last_render = 0.0

    def render_inventory(force: bool = False) -> None:
        nonlocal last_render

        now = time.monotonic()
        if not force and now - last_render < VERTICAL_LOG_INTERVAL:
            return

        last_render = now
        elapsed = max(now - started_at, 0.001)
        speed = completed_skus / elapsed

        print(
            "\n" + "=" * 58 +
            "\nЗАГРУЗКА ОСТАТКОВ" +
            "\n" + "-" * 58 +
            f"\nSKU обработано : {completed_skus} / {total_skus}" +
            f"\nПартии         : {completed_batches} / {len(batches)}" +
            f"\nОтветов        : {len(inventory)}" +
            f"\nHTTP повторы   : {HTTP_RETRIES}" +
            f"\nHTTP ошибки    : {HTTP_FINAL_ERRORS}" +
            f"\nСкорость       : {speed:.1f} SKU/с" +
            "\n" + "=" * 58,
            flush=True,
        )


    print("\nЗагрузка остатков:")
    render_inventory(force=True)

    tasks: list[tuple[int, asyncio.Task[dict[str, dict[str, Any]]]]] = []

    for number, batch in enumerate(batches, start=1):
        task = asyncio.create_task(
            fetch_inventory_batch(
                session,
                semaphore,
                batch,
                number,
                len(batches),
            )
        )
        tasks.append((len(batch), task))

    task_sizes = {
        task: batch_size
        for batch_size, task in tasks
    }

    for future in asyncio.as_completed([task for _, task in tasks]):
        batch_size = task_sizes.get(future, INVENTORY_BATCH_SIZE)

        try:
            batch_result = await future
        except Exception:
            batch_result = {}

        completed_batches += 1
        completed_skus = min(
            total_skus,
            completed_skus + batch_size,
        )

        if isinstance(batch_result, dict):
            inventory.update(batch_result)

        render_inventory(force=True)

    completed_skus = total_skus
    render_inventory(force=True)
    DASHBOARD_ACTIVE = False

    return inventory


def format_excel(file_path: Path) -> None:
    workbook = load_workbook(file_path)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    preferred_widths = {
        "sku": 16,
        "ean": 18,
        "name": 55,
        "brand": 22,
        "price": 13,
        "ordinary_price": 15,
        "regular_price": 15,
        "discount": 13,
        "unit_price": 13,
        "unit": 10,
        "currency": 10,
        "web_qty": 12,
        "web_status": 12,
        "stock_101": 12,
        "stock_103": 12,
        "physical_total": 15,
        "category_level_0": 25,
        "category_level_1": 30,
        "category_level_2": 35,
        "category_path": 60,
        "image_url": 65,
        "product_url": 65,
        "source_category_url": 65,
    }

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        headers = {}

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            headers[str(cell.value)] = cell.column

        for header, column_number in headers.items():
            letter = get_column_letter(column_number)
            worksheet.column_dimensions[letter].width = (
                preferred_widths.get(header, 18)
            )

    workbook.save(file_path)


def format_duration(seconds: float) -> str:
    seconds = max(0, int(round(seconds)))
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def save_json_atomic(path: Path, data: Any) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporary.replace(path)


def save_catalog_cache(products: list[dict[str, Any]], metadata: dict[str, Any]) -> None:
    save_json_atomic(
        CATALOG_DATA_FILE,
        {
            "version": 1,
            "saved_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            "metadata": metadata,
            "products": products,
        },
    )


def load_catalog_cache() -> tuple[list[dict[str, Any]], dict[str, Any]] | None:
    if not CATALOG_DATA_FILE.exists():
        return None
    try:
        data = json.loads(CATALOG_DATA_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    if not isinstance(data, dict) or not isinstance(data.get("products"), list):
        return None
    products = [item for item in data["products"] if isinstance(item, dict)]
    metadata = data.get("metadata") if isinstance(data.get("metadata"), dict) else {}
    return products, metadata


async def collect_full_catalog(
        catalog_session: aiohttp.ClientSession,
        page_semaphore: AdjustableLimiter,
        categories: list[str],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Этап 1: получает все категории, все страницы и сохраняет товарные строки в JSON."""
    global DASHBOARD_ACTIVE, EMPTY_CATEGORY_PAGES
    DASHBOARD_ACTIVE = True

    products: list[dict[str, Any]] = []
    pages_done = 0
    pages_total_known = 0
    categories_done = 0
    current_category = 0
    last_render = 0.0
    stage_started = time.monotonic()

    def render(force: bool = False, status: str = "Сбор каталога") -> None:
        nonlocal last_render
        now = time.monotonic()
        if not force and now - last_render < VERTICAL_LOG_INTERVAL:
            return
        last_render = now
        total_pages = str(pages_total_known) if pages_total_known else "?"
        print(
            "\n" + "=" * 58 +
            "\nЭТАП 1 — СБОР ВСЕГО КАТАЛОГА" +
            "\n" + "-" * 58 +
            f"\nКатегория сейчас: {current_category} / {len(categories)}" +
            f"\nКатегории готовы: {categories_done} / {len(categories)}" +
            f"\nСтраницы       : {pages_done} / {total_pages}" +
            f"\nТоварных строк : {len(products)}" +
            f"\nPAGE limit     : {page_semaphore.limit}" +
            f"\nПустых страниц : {EMPTY_CATEGORY_PAGES}" +
            f"\nHTTP повторы   : {HTTP_RETRIES}" +
            f"\nHTTP ошибки    : {HTTP_FINAL_ERRORS}" +
            f"\nПрошло времени : {format_duration(now - stage_started)}" +
            f"\nСтатус         : {status}" +
            "\n" + "=" * 58,
            flush=True,
        )

    async def accept_hits(hits: list[dict[str, Any]], category_url: str) -> None:
        for hit in hits:
            product = product_from_hit(hit, category_url)
            if product:
                products.append(product)

    async def load_category(category_url: str) -> None:
        global EMPTY_CATEGORY_PAGES
        nonlocal pages_done, pages_total_known
        _, first_hits, metadata = await fetch_category_page(
            catalog_session, page_semaphore, category_url, 1
        )
        if not first_hits:
            await asyncio.sleep(random.uniform(1.0, 2.0))
            _, first_hits, metadata = await fetch_category_page(
                catalog_session, page_semaphore, category_url, 1
            )
        if not first_hits:
            EMPTY_CATEGORY_PAGES += 1
            return

        hits_per_page = metadata["hits_per_page"] or len(first_hits) or 40
        nb_hits = metadata["nb_hits"] or len(first_hits)
        reported_pages = metadata["nb_pages"] or 1
        calculated_pages = max(1, math.ceil(nb_hits / hits_per_page))
        expected_pages = min(
            reported_pages if reported_pages > 0 else calculated_pages,
            MAX_PAGES_PER_CATEGORY,
        )
        pages_total_known += expected_pages
        pages_done += 1
        await accept_hits(first_hits, category_url)
        render()

        tasks = [
            asyncio.create_task(
                fetch_category_page(
                    catalog_session,
                    page_semaphore,
                    category_url,
                    page_number,
                    retry_if_empty=True,
                )
            )
            for page_number in range(2, expected_pages + 1)
        ]
        for future in asyncio.as_completed(tasks):
            _, hits, _ = await future
            pages_done += 1
            if hits:
                await accept_hits(hits, category_url)
            else:
                EMPTY_CATEGORY_PAGES += 1
            render()

    print(
        "\nЗапущен первый этап. Сначала будут загружены абсолютно все "
        "категории, страницы и товарные строки. EAN пока не запускается.\n"
    )
    render(force=True)

    for number, category_url in enumerate(categories, start=1):
        current_category = number
        render(force=True, status=f"Загрузка категории {number}")
        await load_category(category_url)
        categories_done += 1
        render(force=True, status=f"Категория {number} загружена")

    elapsed = time.monotonic() - stage_started
    metadata = {
        "categories": len(categories),
        "pages": pages_done,
        "product_rows": len(products),
        "empty_pages": EMPTY_CATEGORY_PAGES,
        "elapsed_seconds": elapsed,
        "elapsed": format_duration(elapsed),
    }
    await asyncio.to_thread(save_catalog_cache, products, metadata)
    render(force=True, status=f"Каталог сохранён: {CATALOG_DATA_FILE}")
    DASHBOARD_ACTIVE = False
    print(
        f"\nКаталог полностью собран за {format_duration(elapsed)}.\n"
        f"JSON: {CATALOG_DATA_FILE.resolve()}\n"
        f"Товарных строк: {len(products)}"
    )
    return products, metadata


async def enrich_all_products_with_ean(
        ean_session: CurlAsyncSession,
        page_semaphore: AdjustableLimiter,
        ean_semaphore: AdjustableLimiter,
        products: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Этап 2: когда весь каталог уже на руках, получает EAN для всех уникальных SKU."""
    global DASHBOARD_ACTIVE
    DASHBOARD_ACTIVE = True

    cache = load_ean_cache()
    targets: list[tuple[str, str]] = []
    target_seen: set[str] = set()

    # EAN из каталога сразу переносим в кэш. Остальные уникальные SKU ставим в очередь.
    for product in products:
        sku = clean_sku(product.get("sku"))
        url = clean_text(product.get("product_url"))
        catalog_ean = normalize_barcode_candidate(product.get("ean"))
        if not sku:
            continue
        if catalog_ean:
            cache[sku] = {"ean": catalog_ean, "source": "catalog", "url": url}
            continue
        cached = cache.get(sku, {})
        # Наличие source означает, что этот SKU уже проверялся, включая not_found.
        if normalize_barcode_candidate(cached.get("ean")) or clean_text(cached.get("source")):
            continue
        if url and sku not in target_seen:
            target_seen.add(sku)
            targets.append((sku, url))

    ean_queue: asyncio.Queue[tuple[str, str] | None] = asyncio.Queue()
    for item in targets:
        ean_queue.put_nowait(item)

    controller = AdaptiveLoadController(ean_queue, page_semaphore, ean_semaphore)
    await controller.begin_category_ean()

    checked_now = 0
    found_now = 0
    not_found_now = 0
    state_lock = asyncio.Lock()
    speed_window: deque[tuple[float, int]] = deque()
    last_render = 0.0
    stage_started = time.monotonic()
    max_speed = 0.0

    def ean_speed(now: float) -> float:
        cutoff = now - 10.0
        while speed_window and speed_window[0][0] < cutoff:
            speed_window.popleft()
        if len(speed_window) < 2:
            return 0.0
        dt = speed_window[-1][0] - speed_window[0][0]
        if dt <= 0:
            return 0.0
        return (speed_window[-1][1] - speed_window[0][1]) / dt

    def render(force: bool = False, status: str = "Получение EAN") -> None:
        nonlocal last_render, max_speed
        now = time.monotonic()
        if not force and now - last_render < VERTICAL_LOG_INTERVAL:
            return
        last_render = now
        speed = ean_speed(now)
        max_speed = max(max_speed, speed)
        controller.update_speed(speed, checked_now)
        print(
            "\n" + "=" * 58 +
            "\nЭТАП 2 — EAN ДЛЯ ВСЕГО ГОТОВОГО КАТАЛОГА" +
            "\n" + "-" * 58 +
            f"\nТоварных строк : {len(products)}" +
            f"\nУникальных EAN : {len(targets)}" +
            f"\nEAN проверено  : {checked_now} / {len(targets)}" +
            f"\nEAN найдено    : {found_now}" +
            f"\nEAN не найдено : {not_found_now}" +
            f"\nОчередь EAN    : {ean_queue.qsize()}" +
            f"\nСкорость       : {speed:.1f} EAN/с" +
            f"\nEAN limit      : {ean_semaphore.limit}" +
            f"\nEAN активно    : {ean_semaphore.active}" +
            f"\nHTTP 4xx       : {controller.restrictions}" +
            f"\nПауз скорости  : {controller.speed_pauses}" +
            f"\nПрошло времени : {format_duration(now - stage_started)}" +
            f"\nСтатус         : {status}" +
            "\n" + "=" * 58,
            flush=True,
        )

    async def worker() -> None:
        nonlocal checked_now, found_now, not_found_now
        while True:
            item = await ean_queue.get()
            if item is None:
                ean_queue.task_done()
                return
            sku, product_url = item
            try:
                await controller.wait_ean_allowed()
                _, ean, source, resolved_url = await fetch_product_ean(
                    ean_session, ean_semaphore, controller, sku, product_url
                )
            except Exception as error:
                ean = ""
                source = f"worker_error:{type(error).__name__}"
                resolved_url = product_url

            snapshot = None
            async with state_lock:
                cache[sku] = {
                    "ean": ean,
                    "source": source or "not_found",
                    "url": resolved_url,
                }
                checked_now += 1
                speed_window.append((time.monotonic(), checked_now))
                if ean:
                    found_now += 1
                else:
                    not_found_now += 1
                if checked_now % EAN_SAVE_EVERY == 0:
                    snapshot = dict(cache)
                render()
            if snapshot is not None:
                await asyncio.to_thread(save_ean_cache, snapshot)
            ean_queue.task_done()

    print(
        "\nВесь каталог уже находится в JSON/памяти. Запускаем только EAN.\n"
        f"Стартовый limit: {EAN_CONCURRENCY}; работников: {EAN_WORKER_COUNT}; "
        f"товаров для проверки: {len(targets)}.\n"
    )
    render(force=True)

    monitor_task = asyncio.create_task(controller.monitor())
    workers = [asyncio.create_task(worker()) for _ in range(EAN_WORKER_COUNT)]
    await ean_queue.join()

    for _ in workers:
        await ean_queue.put(None)
    await asyncio.gather(*workers, return_exceptions=True)
    await asyncio.to_thread(save_ean_cache, dict(cache))
    monitor_task.cancel()
    await asyncio.gather(monitor_task, return_exceptions=True)

    for product in products:
        sku = clean_sku(product.get("sku"))
        ean = normalize_barcode_candidate(cache.get(sku, {}).get("ean"))
        if ean:
            product["ean"] = ean

    elapsed = time.monotonic() - stage_started
    average_speed = checked_now / elapsed if elapsed > 0 else 0.0
    result = {
        "scheduled": len(targets),
        "checked_now": checked_now,
        "found_now": found_now,
        "not_found_now": not_found_now,
        "rows_filled": sum(bool(normalize_barcode_candidate(p.get("ean"))) for p in products),
        "adaptive_restrictions": controller.restrictions,
        "adaptive_changes": controller.adaptations,
        "speed_pauses": controller.speed_pauses,
        "final_ean_limit": ean_semaphore.limit,
        "average_speed": average_speed,
        "max_speed": max_speed,
        "elapsed_seconds": elapsed,
        "elapsed": format_duration(elapsed),
    }
    render(force=True, status="Все EAN завершены")
    DASHBOARD_ACTIVE = False
    print(f"\nEAN-этап завершён за {format_duration(elapsed)}.")
    return products, result



# ============================================================
# УПРОЩЁННЫЙ ЭКСПОРТ И ЗАПУСК
# ============================================================

OUTPUT_FILE = Path("bauhaus.xlsx")

FINAL_COLUMNS = [
    "Название товара",
    "Цена",
    "Цена со скидкой",
    "Цена со скидкой 2",
    "Штрихкод",
    "Код магазина",
    "Фото",
    "Ссылка",
    "SKU",
    "Category",
    "Category ID",
    "Description",
    "Brand",
    "Model",
]

COLUMNS = FINAL_COLUMNS


class CallbackWriter(io.TextIOBase):
    def __init__(self, callback):
        self.callback = callback
        self._buffer = ""

    def write(self, text):
        if not text:
            return 0
        self._buffer += text
        while "\n" in self._buffer:
            line, self._buffer = self._buffer.split("\n", 1)
            line = line.strip()
            if line and self.callback:
                self.callback(line)
        return len(text)

    def flush(self):
        line = self._buffer.strip()
        if line and self.callback:
            self.callback(line)
        self._buffer = ""


# В упрощённой версии промежуточные JSON-файлы не используются.
def save_catalog_cache(products: list[dict[str, Any]], metadata: dict[str, Any]) -> None:
    return None


def save_ean_cache(cache: dict[str, dict[str, str]]) -> None:
    return None


def load_ean_cache() -> dict[str, dict[str, str]]:
    return {}


def choose_prices(product: dict[str, Any]) -> tuple[float | str, float | str, float | str]:
    final_price = as_number(product.get("price"))
    ordinary_price = as_number(product.get("ordinary_price"))
    regular_price = as_number(product.get("regular_price"))

    base_price = ordinary_price or regular_price or final_price
    sale_price: float | str = ""
    sale_price_2: float | str = ""

    if final_price is not None and base_price is not None and final_price < base_price:
        sale_price = round(final_price, 2)
    elif final_price is not None and base_price is None:
        base_price = final_price

    # Если BAUHAUS отдаёт одновременно ordinary_price и regular_price,
    # сохраняем дополнительный уровень цены только когда он действительно ниже базовой.
    if (
        ordinary_price is not None
        and regular_price is not None
        and regular_price < ordinary_price
        and (sale_price == "" or regular_price != sale_price)
    ):
        base_price = ordinary_price
        sale_price_2 = round(regular_price, 2)

    return (
        round(base_price, 2) if base_price is not None else "",
        sale_price,
        sale_price_2,
    )


def export_final_excel(products: list[dict[str, Any]]) -> None:
    rows: list[dict[str, Any]] = []

    for product in products:
        sku = clean_sku(product.get("sku"))
        name = clean_text(product.get("name"))
        product_url = clean_text(product.get("product_url"))

        if not (sku or name or product_url):
            continue

        price, sale_price, sale_price_2 = choose_prices(product)

        rows.append(
            {
                "Название товара": name,
                "Цена": price,
                "Цена со скидкой": sale_price,
                "Цена со скидкой 2": sale_price_2,
                "Штрихкод": normalize_barcode_candidate(product.get("ean")),
                "Код магазина": sku,
                "Фото": clean_text(
                    product.get("image_url")
                    or product.get("retina_image_url")
                ),
                "Ссылка": product_url,
                "SKU": sku,
                "Category": clean_text(
                    product.get("category_level_2")
                    or product.get("category_level_1")
                    or product.get("category_level_0")
                ),
                "Category ID": "",
                "Description": clean_text(
                    product.get("description")
                    or product.get("short_description")
                ),
                "Brand": clean_text(product.get("brand")),
                "Model": clean_text(product.get("model")),
            }
        )

    dataframe = pd.DataFrame(rows, columns=FINAL_COLUMNS).fillna("")

    if dataframe.empty:
        raise RuntimeError("Нет товаров для сохранения.")

    with_sku = dataframe[dataframe["Код магазина"] != ""].drop_duplicates(
        subset=["Код магазина"], keep="first"
    )
    without_sku = dataframe[dataframe["Код магазина"] == ""].drop_duplicates(
        subset=["Ссылка", "Название товара"], keep="first"
    )

    dataframe = pd.concat([with_sku, without_sku], ignore_index=True)
    dataframe = dataframe.sort_values(
        by=["Название товара", "Код магазина"],
        na_position="last",
    ).reset_index(drop=True)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        dataframe.to_excel(writer, sheet_name="Товары", index=False)
        worksheet = writer.sheets["Товары"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        widths = {
            "A": 65,
            "B": 14,
            "C": 18,
            "D": 20,
            "E": 22,
            "F": 20,
            "G": 70,
            "H": 80,
        }
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width

        for row_number in range(2, worksheet.max_row + 1):
            worksheet[f"E{row_number}"].number_format = "@"
            worksheet[f"F{row_number}"].number_format = "@"
            for column in ("B", "C", "D"):
                worksheet[f"{column}{row_number}"].number_format = "0.00"

    print(f"\nГотово: {OUTPUT_FILE.resolve()}")
    print(f"Строк в Excel: {len(dataframe)}")


async def _main() -> None:
    catalog_connector = aiohttp.TCPConnector(
        limit=20,
        limit_per_host=12,
        ttl_dns_cache=600,
        keepalive_timeout=30,
        enable_cleanup_closed=True,
    )

    page_semaphore = AdjustableLimiter(PAGE_CONCURRENCY)

    async with aiohttp.ClientSession(
        headers=HEADERS,
        connector=catalog_connector,
        cookie_jar=aiohttp.CookieJar(),
    ) as catalog_session:
        print("Открываем BAUHAUS и получаем дерево категорий...")

        home = await request_text(
            catalog_session,
            BASE_URL,
            page_semaphore,
        )

        if not home:
            raise RuntimeError("Не удалось открыть главную страницу BAUHAUS.")

        categories = await discover_top_level_categories(
            catalog_session,
            page_semaphore,
            home,
        )

        if not categories:
            raise RuntimeError("Не удалось найти категории с товарами.")

        products, _ = await collect_full_catalog(
            catalog_session,
            page_semaphore,
            categories,
        )

        if not products:
            raise RuntimeError("Каталог BAUHAUS не вернул товары.")

    # EAN enrichment is intentionally not part of the primary Excel Pipeline.
    # Product pages can be slow or unavailable, while an empty barcode is valid.
    await asyncio.to_thread(export_final_excel, products)
    print(
        "Основной каталог BAUHAUS сохранён. "
        "EAN пропущен и может быть получен отдельным необязательным этапом."
    )


async def main(output_path: str | Path | None = None, log_callback=None) -> None:
    global OUTPUT_FILE

    original_output_file = OUTPUT_FILE
    if output_path is not None:
        OUTPUT_FILE = Path(output_path)
        OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    writer = CallbackWriter(log_callback)
    try:
        if log_callback is None:
            await _main()
        else:
            with contextlib.redirect_stdout(writer):
                await _main()
            writer.flush()
    finally:
        OUTPUT_FILE = original_output_file


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\nПарсер остановлен пользователем.")
    except Exception as error:
        print(
            f"\nКритическая ошибка: "
            f"{type(error).__name__}: {error}"
        )
        raise
