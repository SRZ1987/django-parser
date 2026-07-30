from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


@dataclass
class ExcelValidationResult:
    is_valid: bool
    rows_count: int = 0
    error_message: str = ""


class ExcelValidationError(Exception):
    pass


class ExcelCatalogValidator:
    def validate(self, path, *, column_map, worksheet_name=None):
        path = Path(path)
        if not path.exists():
            return ExcelValidationResult(False, error_message="Excel file does not exist.")

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            return ExcelValidationResult(False, error_message=f"Excel file cannot be opened: {exc}")

        try:
            worksheet = workbook[worksheet_name] if worksheet_name else workbook.active
            headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
            missing = [column for column in column_map if column not in headers]
            if missing:
                return ExcelValidationResult(False, error_message=f"Missing required columns: {', '.join(missing)}")

            rows_count = 0
            header_index = {header: index for index, header in enumerate(headers)}
            price_columns = [column for column, target in column_map.items() if target in {"price", "sale_price"}]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if not any(value not in (None, "") for value in row):
                    continue
                rows_count += 1
                for column in price_columns:
                    value = row[header_index[column]]
                    if value in (None, ""):
                        continue
                    if parse_decimal(value) is None:
                        return ExcelValidationResult(False, error_message=f"Invalid price in column {column}.")

            if rows_count <= 0:
                return ExcelValidationResult(False, error_message="Excel catalog is empty.")
            return ExcelValidationResult(True, rows_count=rows_count)
        finally:
            workbook.close()


def parse_decimal(value):
    if value in (None, "") or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value
    try:
        parsed = Decimal(str(value).replace(",", ".").replace(" ", "")).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None
    return parsed if parsed >= 0 else None
