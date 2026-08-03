import hashlib
import time
from dataclasses import dataclass, field

from django.core.exceptions import ValidationError
from django.db import DataError, IntegrityError, transaction
from django.utils import timezone
from openpyxl import load_workbook

from catalog.models import PriceHistory, Product, ProductOffer
from catalog.services.normalization import normalize_product_name

from .excel_validation import parse_decimal


class ExcelImportError(ValueError):
    pass


class ExcelFieldLengthError(ExcelImportError):
    pass


@dataclass
class ExcelRowError:
    row_number: int
    reason: str
    recoverable: bool = False

    def format(self):
        return f"row {self.row_number}: {self.reason}"


@dataclass
class ExcelImportResult:
    total_rows: int = 0
    valid_rows: int = 0
    products_found: int = 0
    products_created: int = 0
    products_updated: int = 0
    prices_changed: int = 0
    skipped_rows: int = 0
    errors_count: int = 0
    row_errors: list[ExcelRowError] = field(default_factory=list)
    created_offer_ids: list[int] = field(default_factory=list)


class ExcelCatalogImporter:
    DEACTIVATE_BATCH_SIZE = 1000
    MIN_ACTIVE_OFFERS_FOR_ANOMALY_CHECK = 100
    MIN_REMOTE_TO_ACTIVE_RATIO = 0.2
    MAX_INVALID_ROW_RATIO = 0.05
    HEARTBEAT_INTERVAL_SECONDS = 15

    def import_file(self, parser_export, *, column_map, worksheet_name=None, deactivate_missing=True, parser_run=None):
        result, parsed_rows, remote_external_ids = self._read_and_normalize(parser_export, column_map, worksheet_name, parser_run)

        try:
            can_deactivate = deactivate_missing and not result.row_errors
            self._validate_import_readiness(parser_export.shop, result, remote_external_ids, can_deactivate)
            self._check_cancel_requested(parser_run, "Parser run was cancelled before database import.")

            seen_at = timezone.now()
            last_heartbeat = time.monotonic()
            database_row_errors = []
            with transaction.atomic():
                for parsed in parsed_rows:
                    try:
                        created, price_changed, offer_id = self._save_offer(parser_export.shop, parsed, seen_at)
                    except (DataError, IntegrityError, ValidationError, ValueError) as exc:
                        row_error = self._add_row_issue(
                            result,
                            parsed["_row_number"],
                            f"database save failed: {type(exc).__name__}: {exc}",
                            skipped=True,
                            error=True,
                            recoverable=True,
                        )
                        database_row_errors.append(row_error)
                    else:
                        result.products_found += 1
                        result.products_created += int(created)
                        result.products_updated += int(not created)
                        result.prices_changed += int(price_changed)
                        if created:
                            result.created_offer_ids.append(offer_id)
                    last_heartbeat = self._heartbeat(
                        parser_run,
                        last_heartbeat,
                        f"Imported {result.products_found}/{result.valid_rows} Excel rows.",
                    )

                if database_row_errors:
                    self._log_row_errors(parser_run, database_row_errors)
                    database_error_ratio = len(database_row_errors) / max(result.valid_rows, 1)
                    if len(database_row_errors) > 1 and database_error_ratio > self.MAX_INVALID_ROW_RATIO:
                        raise ExcelImportError(
                            f"Database import has too many row errors: {len(database_row_errors)}/"
                            f"{result.valid_rows} ({database_error_ratio:.1%})."
                        )
                if result.products_found <= 0:
                    raise ExcelImportError("Excel import did not save any product rows.")

                can_deactivate = deactivate_missing and not result.row_errors
                if can_deactivate:
                    self._check_cancel_requested(parser_run, "Parser run was cancelled before deactivation.")
                    self._deactivate_missing_offers(parser_export.shop, remote_external_ids)

            parser_export.imported_at = timezone.now()
            parser_export.import_success = True
            parser_export.validation_error = (
                self._format_validation_error("Excel imported with skipped rows.", result.row_errors)
                if result.row_errors
                else ""
            )
            parser_export.save(update_fields=["imported_at", "import_success", "validation_error"])
            return result
        except Exception as exc:
            parser_export.import_success = False
            parser_export.validation_error = self._format_validation_error(str(exc), result.row_errors)
            parser_export.save(update_fields=["import_success", "validation_error"])
            raise

    def _read_and_normalize(self, parser_export, column_map, worksheet_name, parser_run):
        result = ExcelImportResult()
        parsed_rows = []
        remote_external_ids = set()
        last_heartbeat = time.monotonic()

        with parser_export.file.open("rb") as file_obj:
            workbook = load_workbook(file_obj, read_only=True, data_only=True)
            try:
                worksheet = workbook[worksheet_name] if worksheet_name else workbook.active
                headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
                header_index = {header: index for index, header in enumerate(headers)}
                self._validate_headers(header_index, column_map)

                for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(value not in (None, "") for value in row):
                        continue
                    result.total_rows += 1
                    try:
                        parsed = self._parse_row(row, header_index, column_map)
                        if not parsed["external_id"]:
                            self._add_row_issue(result, row_number, "missing external_id", skipped=True)
                            continue
                        if not parsed["original_name"]:
                            self._add_row_issue(result, row_number, "missing original_name", skipped=True)
                            continue
                        self._validate_model_field_lengths(parsed)
                        parsed["_row_number"] = row_number
                        parsed_rows.append(parsed)
                        remote_external_ids.add(parsed["external_id"])
                        result.valid_rows += 1
                    except ExcelFieldLengthError as exc:
                        self._add_row_issue(
                            result,
                            row_number,
                            str(exc),
                            skipped=True,
                            error=True,
                            recoverable=True,
                        )
                    except Exception as exc:
                        self._add_row_issue(result, row_number, str(exc), skipped=True, error=True)

                    last_heartbeat = self._heartbeat(
                        parser_run,
                        last_heartbeat,
                        f"Read {result.total_rows} Excel rows; valid={result.valid_rows}; errors={result.errors_count}.",
                    )
            finally:
                workbook.close()

        self._log_row_errors(parser_run, result.row_errors)
        return result, parsed_rows, remote_external_ids

    def _validate_headers(self, header_index, column_map):
        missing_columns = [column for column in column_map if column not in header_index]
        if missing_columns:
            raise ExcelImportError(f"Excel is missing required columns: {', '.join(missing_columns)}")

    def _validate_import_readiness(self, shop, result, remote_external_ids, deactivate_missing):
        if result.valid_rows <= 0:
            raise ExcelImportError("Excel import produced no valid product rows.")

        invalid_rows = sum(not row_error.recoverable for row_error in result.row_errors)
        invalid_ratio = invalid_rows / max(result.total_rows, 1)
        if invalid_ratio > self.MAX_INVALID_ROW_RATIO:
            raise ExcelImportError(
                f"Excel import has too many invalid rows: {invalid_rows}/{result.total_rows} "
                f"({invalid_ratio:.1%})."
            )

        recoverable_rows = sum(row_error.recoverable for row_error in result.row_errors)
        recoverable_ratio = recoverable_rows / max(result.total_rows, 1)
        if recoverable_rows > 1 and recoverable_ratio > self.MAX_INVALID_ROW_RATIO:
            raise ExcelImportError(
                f"Excel import has too many recoverable row errors: {recoverable_rows}/{result.total_rows} "
                f"({recoverable_ratio:.1%})."
            )

        if deactivate_missing:
            self._validate_catalog_size(shop, remote_external_ids)

    def _parse_row(self, row, header_index, column_map):
        data = {}
        for column, target in column_map.items():
            index = header_index[column]
            data[target] = clean_value(row[index] if index < len(row) else "")

        external_id = data.get("external_id") or stable_external_id(data.get("product_url"), data.get("original_name"))
        price = parse_decimal(data.get("price"))
        sale_price = parse_decimal(data.get("sale_price"))
        if price is None and sale_price is not None:
            price = sale_price
            sale_price = None
        if price is not None and sale_price is not None and sale_price >= price:
            sale_price = None

        return {
            "external_id": external_id,
            "sku": data.get("external_id", ""),
            "barcode": data.get("barcode", ""),
            "original_name": data.get("original_name", ""),
            "price": price,
            "sale_price": sale_price,
            "product_url": data.get("product_url", ""),
            "image_url": data.get("image_url", ""),
        }

    def _validate_model_field_lengths(self, parsed):
        normalized_name = normalize_product_name(parsed["original_name"])
        fields = (
            ("external_id", parsed["external_id"], ProductOffer, "external_id"),
            ("sku", parsed["sku"], ProductOffer, "sku"),
            ("barcode", parsed["barcode"], ProductOffer, "barcode"),
            ("barcode", parsed["barcode"], Product, "barcode"),
            ("original_name", parsed["original_name"], ProductOffer, "original_name"),
            ("name", parsed["original_name"], Product, "name"),
            ("normalized_name", normalized_name, ProductOffer, "normalized_name"),
            ("normalized_name", normalized_name, Product, "normalized_name"),
            ("product_url", parsed["product_url"], ProductOffer, "product_url"),
            ("image_url", parsed["image_url"], ProductOffer, "image_url"),
        )
        for source_field, value, model, model_field_name in fields:
            model_field = model._meta.get_field(model_field_name)
            max_length = model_field.max_length
            if max_length is None or value is None:
                continue
            value = str(value)
            if len(value) <= max_length:
                continue
            raise ExcelFieldLengthError(
                f"field={model.__name__}.{model_field.name} source={source_field} "
                f"max_length={max_length} actual_length={len(value)} "
                f"value_preview={value[:200]!r}"
            )

    @transaction.atomic
    def _save_offer(self, shop, parsed, seen_at):
        offer = ProductOffer.objects.select_related("product").filter(shop=shop, external_id=parsed["external_id"]).first()
        created = offer is None
        if created:
            product = Product.objects.create(
                name=parsed["original_name"],
                barcode=parsed["barcode"],
                normalized_name=normalize_product_name(parsed["original_name"]),
            )
            offer = ProductOffer(shop=shop, product=product, external_id=parsed["external_id"])
        else:
            product = offer.product
            product.name = parsed["original_name"]
            if parsed["barcode"]:
                product.barcode = parsed["barcode"]
            product.normalized_name = normalize_product_name(parsed["original_name"])
            product.save(update_fields=["name", "normalized_name", "barcode", "updated_at"])

        previous_price = offer.price
        previous_sale_price = offer.sale_price
        offer.original_name = parsed["original_name"]
        offer.normalized_name = normalize_product_name(parsed["original_name"])
        offer.sku = parsed["sku"]
        if parsed["barcode"]:
            offer.barcode = parsed["barcode"]
        offer.price = parsed["price"] if parsed["price"] is not None else offer.price
        offer.sale_price = parsed["sale_price"]
        offer.currency = "EUR"
        if parsed["product_url"]:
            offer.product_url = parsed["product_url"]
        if parsed["image_url"]:
            offer.image_url = parsed["image_url"]
        offer.is_active = True
        offer.is_available = True
        offer.last_seen_at = seen_at
        offer.save()

        price_changed = previous_price != offer.price or previous_sale_price != offer.sale_price
        if (created and (offer.price is not None or offer.sale_price is not None)) or price_changed:
            PriceHistory.objects.create(offer=offer, price=offer.price, sale_price=offer.sale_price)
            return created, True, offer.pk
        return created, False, offer.pk

    def _validate_catalog_size(self, shop, remote_external_ids):
        active_count = ProductOffer.objects.filter(shop=shop, is_active=True).count()
        if active_count < self.MIN_ACTIVE_OFFERS_FOR_ANOMALY_CHECK:
            return
        if len(remote_external_ids) >= active_count * self.MIN_REMOTE_TO_ACTIVE_RATIO:
            return
        raise ExcelImportError(
            f"Excel import returned an anomalously small product list: {len(remote_external_ids)} "
            f"remote products for {active_count} active offers."
        )

    def _deactivate_missing_offers(self, shop, remote_external_ids):
        missing_offer_ids = [
            offer_id
            for offer_id, external_id in ProductOffer.objects.filter(shop=shop).values_list("id", "external_id")
            if external_id not in remote_external_ids
        ]
        for start in range(0, len(missing_offer_ids), self.DEACTIVATE_BATCH_SIZE):
            ProductOffer.objects.filter(id__in=missing_offer_ids[start : start + self.DEACTIVATE_BATCH_SIZE]).update(
                is_active=False,
                is_available=False,
            )

    def _add_row_issue(
        self,
        result,
        row_number,
        reason,
        *,
        skipped=False,
        error=False,
        recoverable=False,
    ):
        result.skipped_rows += int(skipped)
        result.errors_count += int(error)
        row_error = ExcelRowError(row_number=row_number, reason=reason, recoverable=recoverable)
        result.row_errors.append(row_error)
        return row_error

    def _format_validation_error(self, message, row_errors):
        details = "\n".join(error.format() for error in row_errors[:50])
        return f"{message}\n{details}" if details else message

    def _log_row_errors(self, parser_run, row_errors):
        if parser_run is None or not row_errors:
            return
        message = self._format_validation_error("Excel row errors:", row_errors)
        parser_run.log = f"{parser_run.log}\n{message}" if parser_run.log else message
        parser_run.heartbeat_at = timezone.now()
        parser_run.save(update_fields=["log", "heartbeat_at"])

    def _heartbeat(self, parser_run, last_heartbeat, message):
        if parser_run is None:
            return last_heartbeat
        now = time.monotonic()
        if now - last_heartbeat < self.HEARTBEAT_INTERVAL_SECONDS:
            return last_heartbeat
        parser_run.log = f"{parser_run.log}\n{message}" if parser_run.log else message
        parser_run.heartbeat_at = timezone.now()
        parser_run.save(update_fields=["log", "heartbeat_at"])
        return now

    def _check_cancel_requested(self, parser_run, message):
        if parser_run is None:
            return
        parser_run.refresh_from_db(fields=["cancel_requested"])
        if parser_run.cancel_requested:
            raise ExcelImportError(message)


def clean_value(value):
    return "" if value in (None, "") else str(value).strip()


def stable_external_id(*parts):
    source = "|".join(clean_value(part) for part in parts if clean_value(part))
    return hashlib.sha256(source.encode("utf-8")).hexdigest()[:32] if source else ""
