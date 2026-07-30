import contextlib
import io
from pathlib import Path

from parsers.standalone import espak_parser

from .base import ParserAdapter, ParserResult


class CallbackWriter(io.TextIOBase):
    def __init__(self, callback):
        self.callback = callback
        self._buffer = ""

    def write(self, text):
        if not text:
            return 0
        self._buffer += text
        while "\n" in self._buffer:
            line, self._buffer = self._buffer.split("\n", 1)
            line = line.strip()
            if line and self.callback:
                self.callback(line)
        return len(text)

    def flush(self):
        line = self._buffer.strip()
        if line and self.callback:
            self.callback(line)
        self._buffer = ""


class EspakAdapter(ParserAdapter):
    code = "espak"
    name = "ESPAK"
    worksheet_name = None
    column_map = {
        espak_parser.COLUMNS[0]: "original_name",
        espak_parser.COLUMNS[1]: "price",
        espak_parser.COLUMNS[2]: "sale_price",
        espak_parser.COLUMNS[4]: "barcode",
        espak_parser.COLUMNS[5]: "external_id",
        espak_parser.COLUMNS[6]: "image_url",
        espak_parser.COLUMNS[7]: "product_url",
    }

    async def run(self, output_path, log_callback=None):
        original_output_file = espak_parser.OUTPUT_FILE
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        espak_parser.OUTPUT_FILE = output_path

        writer = CallbackWriter(log_callback)
        try:
            with contextlib.redirect_stdout(writer):
                await espak_parser.main()
            writer.flush()
            return ParserResult(
                success=True,
                output_path=str(output_path),
                products_count=count_excel_rows(output_path),
            )
        except Exception as exc:
            return ParserResult(
                success=False,
                output_path=str(output_path),
                products_count=0,
                error_message=str(exc),
            )
        finally:
            espak_parser.OUTPUT_FILE = original_output_file


def count_excel_rows(path):
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        return max(worksheet.max_row - 1, 0)
    finally:
        workbook.close()
