from openpyxl import load_workbook


def count_excel_rows(path, worksheet_name=None):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[worksheet_name] if worksheet_name else workbook.active
        return max(worksheet.max_row - 1, 0)
    finally:
        workbook.close()
