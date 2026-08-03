import tempfile
import threading
import asyncio
import time
from datetime import timedelta
from concurrent.futures import ThreadPoolExecutor
from io import StringIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock, patch

from django.contrib.auth import get_user_model
from django.core.files import File
from django.core.management import call_command
from django.db import OperationalError
from django.test import Client, TestCase, TransactionTestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from catalog.models import Product, ProductOffer, Shop
from parsers.adapters.base import ParserResult
from parsers.adapters.bauhaus import BauhausAdapter
from parsers.adapters.bauhof import BauhofAdapter
from parsers.adapters.depo import DepoAdapter
from parsers.adapters.ehituseabc import EhituseABCAdapter
from parsers.adapters.espak import EspakAdapter
from parsers.adapters.fere import FereAdapter
from parsers.adapters.handymann import HandymannAdapter
from parsers.adapters.registry import ADAPTERS
from parsers.standalone.public_commerce_parser import PUBLIC_COMMERCE_STORES
from parsers.models import ParserBatch, ParserBatchLock, ParserConfig, ParserExport, ParserQueueJob, ParserRun
from parsers.services.batch_runner import (
    ParserBatchAlreadyRunning,
    ParserBatchLockMissing,
    process_next_queue_job,
    run_all_parsers,
    run_excel_parser,
    start_batch,
)
from parsers.services.export_storage import export_work_paths
from parsers.services.excel_importer import ExcelCatalogImporter, ExcelImportError
from parsers.services.excel_validation import ExcelCatalogValidator
from parsers.services.heartbeat import HeartbeatTicker
from parsers.services.recovery import (
    STALE_BATCH_MESSAGE,
    STALE_JOB_MESSAGE,
    STALE_RUN_MESSAGE,
    recover_stale_parser_state,
)
from parsers.standalone import bauhaus_parser
from parsers.standalone import bauhof_parser
from parsers.standalone import depo_parser
from parsers.standalone import espak_parser
from parsers.standalone import ehituseabc_parser
from parsers.standalone import fere_parser


def create_xlsx(path, headers=None, rows=None, sheet_name=None):
    workbook = Workbook()
    worksheet = workbook.active
    if sheet_name:
        worksheet.title = sheet_name
    worksheet.append(headers or espak_parser.COLUMNS)
    for row in rows or []:
        worksheet.append(row)
    workbook.save(path)


class ExcelValidationTests(TestCase):
    def test_empty_excel_is_invalid(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "empty.xlsx"
            create_xlsx(path)

            result = ExcelCatalogValidator().validate(path, column_map=EspakAdapter.column_map)

        self.assertFalse(result.is_valid)
        self.assertIn("empty", result.error_message.lower())

    def test_corrupted_excel_is_invalid(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "broken.xlsx"
            path.write_text("not an xlsx", encoding="utf-8")

            result = ExcelCatalogValidator().validate(path, column_map=EspakAdapter.column_map)

        self.assertFalse(result.is_valid)
        self.assertIn("cannot be opened", result.error_message)


class ExportStorageTests(TestCase):
    def test_export_work_paths_are_unique_for_same_parser_code(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            with override_settings(PARSER_EXPORT_WORK_DIR=Path(tmp_dir)):
                first_tmp, first_final = export_work_paths("espak")
                second_tmp, second_final = export_work_paths("espak")

        self.assertNotEqual(first_tmp, second_tmp)
        self.assertNotEqual(first_final, second_final)
        self.assertEqual(first_tmp.suffix, ".xlsx")
        self.assertTrue(first_tmp.name.endswith(".tmp.xlsx"))


class ExcelImportTests(TestCase):
    def setUp(self):
        self.shop = Shop.objects.create(name="ESPAK", code="espak")
        self.config = ParserConfig.objects.create(shop=self.shop, name="ESPAK parser", code="espak")
        self.run = ParserRun.objects.create(parser=self.config, trigger=ParserRun.TRIGGER_COMMAND)

    @override_settings(MEDIA_ROOT=tempfile.mkdtemp())
    def test_successful_excel_import_creates_offer(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "espak.xlsx"
            create_xlsx(
                path,
                rows=[["Hammer", 10, 8, "", "4740000000001", "SKU-1", "https://img.test/a.jpg", "https://espak.test/a"]],
            )
            parser_export = self._create_export(path, rows_count=1)

            result = ExcelCatalogImporter().import_file(parser_export, column_map=EspakAdapter.column_map)

        offer = ProductOffer.objects.get(shop=self.shop, external_id="SKU-1")
        self.assertEqual(result.products_created, 1)
        self.assertEqual(offer.original_name, "Hammer")
        self.assertEqual(str(offer.price), "10.00")
        self.assertEqual(str(offer.sale_price), "8.00")
        self.assertEqual(offer.barcode, "4740000000001")

    @override_settings(MEDIA_ROOT=tempfile.mkdtemp())
    def test_depo_quantity_price_is_imported_and_updated_without_duplicates(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            first_path = Path(tmp_dir) / "depo-first.xlsx"
            create_xlsx(
                first_path,
                headers=depo_parser.COLUMNS,
                rows=[
                    [
                        "DEPO construction screw",
                        10,
                        "",
                        8.5,
                        "4740000000001",
                        "DEPO-1",
                        "https://img.test/depo.jpg",
                        "https://online.depo.ee/product/DEPO-1",
                        6,
                    ]
                ],
            )
            first_export = self._create_export(first_path, rows_count=1)
            first_result = ExcelCatalogImporter().import_file(first_export, column_map=DepoAdapter.column_map)

            second_path = Path(tmp_dir) / "depo-second.xlsx"
            create_xlsx(
                second_path,
                headers=depo_parser.COLUMNS,
                rows=[
                    [
                        "DEPO construction screw",
                        10,
                        "",
                        7.75,
                        "4740000000001",
                        "DEPO-1",
                        "https://img.test/depo.jpg",
                        "https://online.depo.ee/product/DEPO-1",
                        10,
                    ]
                ],
            )
            self.run = ParserRun.objects.create(parser=self.config, trigger=ParserRun.TRIGGER_COMMAND)
            second_export = self._create_export(second_path, rows_count=1)
            second_result = ExcelCatalogImporter().import_file(second_export, column_map=DepoAdapter.column_map)

        offer = ProductOffer.objects.get(shop=self.shop, external_id="DEPO-1")
        self.assertEqual(first_result.products_created, 1)
        self.assertEqual(second_result.products_updated, 1)
        self.assertEqual(ProductOffer.objects.filter(shop=self.shop, external_id="DEPO-1").count(), 1)
        self.assertIsNone(offer.sale_price)
        self.assertEqual(str(offer.quantity_price), "7.75")
        self.assertEqual(offer.quantity_price_min_quantity, 10)
        self.assertEqual(offer.price_for_quantity(9), offer.price)
        self.assertEqual(offer.price_for_quantity(10), offer.quantity_price)
        self.assertEqual(offer.price_history.count(), 2)
        latest_history = offer.price_history.first()
        self.assertEqual(str(latest_history.quantity_price), "7.75")
        self.assertEqual(latest_history.quantity_price_min_quantity, 10)

    @override_settings(MEDIA_ROOT=tempfile.mkdtemp())
    def test_overlong_field_reports_row_field_length_and_preview(self):
        long_external_id = str(["PC692956"] + [str(1558800 + index) for index in range(20)])
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "overlong.xlsx"
            create_xlsx(
                path,
                rows=[["Invalid hammer", 10, "", "", "", long_external_id, "", ""]],
            )
            parser_export = self._create_export(path, rows_count=1)

            with self.assertRaises(ExcelImportError):
                ExcelCatalogImporter().import_file(
                    parser_export,
                    column_map=EspakAdapter.column_map,
                    parser_run=self.run,
                )

        parser_export.refresh_from_db()
        self.run.refresh_from_db()
        expected_details = [
            "row 2",
            "field=ProductOffer.external_id",
            "max_length=150",
            f"actual_length={len(long_external_id)}",
            f"value_preview={long_external_id[:200]!r}",
        ]
        for detail in expected_details:
            self.assertIn(detail, parser_export.validation_error)
            self.assertIn(detail, self.run.log)

    @override_settings(MEDIA_ROOT=tempfile.mkdtemp())
    def test_one_overlong_row_does_not_roll_back_valid_rows(self):
        old_product = Product.objects.create(name="Existing hammer")
        old_offer = ProductOffer.objects.create(
            shop=self.shop,
            product=old_product,
            external_id="SKU-OLD",
            sku="SKU-OLD",
            original_name="Existing hammer",
            is_active=True,
            is_available=True,
        )
        long_external_id = "X" * 151
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "partial.xlsx"
            create_xlsx(
                path,
                rows=[
                    ["Invalid hammer", 9, "", "", "", long_external_id, "", ""],
                    ["Valid hammer", 10, "", "", "", "SKU-VALID", "", ""],
                ],
            )
            parser_export = self._create_export(path, rows_count=2)

            result = ExcelCatalogImporter().import_file(
                parser_export,
                column_map=EspakAdapter.column_map,
                parser_run=self.run,
            )

        parser_export.refresh_from_db()
        old_offer.refresh_from_db()
        self.assertEqual(result.products_created, 1)
        self.assertEqual(result.products_found, 1)
        self.assertEqual(result.errors_count, 1)
        self.assertEqual(result.skipped_rows, 1)
        self.assertTrue(ProductOffer.objects.filter(shop=self.shop, external_id="SKU-VALID").exists())
        self.assertFalse(ProductOffer.objects.filter(shop=self.shop, external_id=long_external_id).exists())
        self.assertTrue(old_offer.is_active)
        self.assertTrue(old_offer.is_available)
        self.assertTrue(parser_export.import_success)
        self.assertIn("row 2", parser_export.validation_error)
        self.assertIn("field=ProductOffer.external_id", parser_export.validation_error)

    @override_settings(MEDIA_ROOT=tempfile.mkdtemp())
    def test_one_database_row_error_does_not_roll_back_other_rows(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "database-row-error.xlsx"
            create_xlsx(
                path,
                rows=[
                    ["Broken hammer", 9, "", "", "", "SKU-BROKEN", "", ""],
                    ["Valid hammer", 10, "", "", "", "SKU-VALID", "", ""],
                ],
            )
            parser_export = self._create_export(path, rows_count=2)
            importer = ExcelCatalogImporter()
            save_offer = importer._save_offer

            def save_with_one_error(shop, parsed, seen_at):
                if parsed["external_id"] == "SKU-BROKEN":
                    raise ValueError("invalid row data")
                return save_offer(shop, parsed, seen_at)

            with patch.object(importer, "_save_offer", side_effect=save_with_one_error):
                result = importer.import_file(
                    parser_export,
                    column_map=EspakAdapter.column_map,
                    parser_run=self.run,
                )

        parser_export.refresh_from_db()
        self.assertEqual(result.products_created, 1)
        self.assertEqual(result.products_found, 1)
        self.assertEqual(result.errors_count, 1)
        self.assertEqual(result.skipped_rows, 1)
        self.assertFalse(ProductOffer.objects.filter(shop=self.shop, external_id="SKU-BROKEN").exists())
        self.assertTrue(ProductOffer.objects.filter(shop=self.shop, external_id="SKU-VALID").exists())
        self.assertTrue(parser_export.import_success)
        self.assertIn("row 2", parser_export.validation_error)
        self.assertIn("database save failed: ValueError: invalid row data", parser_export.validation_error)

    @override_settings(MEDIA_ROOT=tempfile.mkdtemp())
    def test_empty_barcode_does_not_overwrite_existing_barcode(self):
        product = Product.objects.create(name="Hammer", barcode="4740000000001")
        ProductOffer.objects.create(
            shop=self.shop,
            product=product,
            external_id="SKU-1",
            sku="SKU-1",
            barcode="4740000000001",
            original_name="Hammer",
            product_url="https://espak.test/a",
            image_url="https://img.test/a.jpg",
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "espak.xlsx"
            create_xlsx(
                path,
                rows=[["Hammer updated", 11, "", "", "", "SKU-1", "", ""]],
            )
            parser_export = self._create_export(path, rows_count=1)

            ExcelCatalogImporter().import_file(parser_export, column_map=EspakAdapter.column_map)

        offer = ProductOffer.objects.get(shop=self.shop, external_id="SKU-1")
        self.assertEqual(offer.barcode, "4740000000001")
        self.assertEqual(offer.product_url, "https://espak.test/a")
        self.assertEqual(offer.image_url, "https://img.test/a.jpg")

    @override_settings(MEDIA_ROOT=tempfile.mkdtemp())
    def test_failed_import_does_not_deactivate_existing_offer(self):
        product = Product.objects.create(name="Hammer")
        ProductOffer.objects.create(
            shop=self.shop,
            product=product,
            external_id="SKU-1",
            sku="SKU-1",
            original_name="Hammer",
            is_active=True,
            is_available=True,
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "espak.xlsx"
            create_xlsx(path, rows=[["", "", "", "", "", "", "", ""]])
            parser_export = self._create_export(path, rows_count=1)

            with self.assertRaises(ValueError):
                ExcelCatalogImporter().import_file(parser_export, column_map=EspakAdapter.column_map)

        offer = ProductOffer.objects.get(shop=self.shop, external_id="SKU-1")
        self.assertTrue(offer.is_active)
        self.assertTrue(offer.is_available)

    @override_settings(MEDIA_ROOT=tempfile.mkdtemp())
    def test_error_after_excel_read_does_not_leave_partial_import(self):
        product = Product.objects.create(name="Old hammer")
        ProductOffer.objects.create(
            shop=self.shop,
            product=product,
            external_id="SKU-OLD",
            sku="SKU-OLD",
            original_name="Old hammer",
            is_active=True,
            is_available=True,
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "espak.xlsx"
            create_xlsx(path, rows=[["New hammer", 10, "", "", "", "SKU-NEW", "", ""]])
            parser_export = self._create_export(path, rows_count=1)

            with patch.object(ExcelCatalogImporter, "_deactivate_missing_offers", side_effect=RuntimeError("boom")):
                with self.assertRaises(RuntimeError):
                    ExcelCatalogImporter().import_file(parser_export, column_map=EspakAdapter.column_map)

        self.assertFalse(ProductOffer.objects.filter(shop=self.shop, external_id="SKU-NEW").exists())
        self.assertTrue(ProductOffer.objects.get(shop=self.shop, external_id="SKU-OLD").is_active)

    @override_settings(MEDIA_ROOT=tempfile.mkdtemp())
    def test_erroneous_rows_do_not_deactivate_missing_offers(self):
        product = Product.objects.create(name="Old hammer")
        ProductOffer.objects.create(
            shop=self.shop,
            product=product,
            external_id="SKU-OLD",
            sku="SKU-OLD",
            original_name="Old hammer",
            is_active=True,
            is_available=True,
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "espak.xlsx"
            create_xlsx(
                path,
                rows=[
                    ["New hammer", 10, "", "", "", "SKU-NEW", "", ""],
                    ["", 12, "", "", "", "SKU-BAD", "", ""],
                ],
            )
            parser_export = self._create_export(path, rows_count=2)

            with self.assertRaises(ExcelImportError):
                ExcelCatalogImporter().import_file(parser_export, column_map=EspakAdapter.column_map)

        parser_export.refresh_from_db()
        old_offer = ProductOffer.objects.get(shop=self.shop, external_id="SKU-OLD")
        self.assertTrue(old_offer.is_active)
        self.assertFalse(parser_export.import_success)
        self.assertIn("row 3", parser_export.validation_error)
        self.assertFalse(ProductOffer.objects.filter(shop=self.shop, external_id="SKU-NEW").exists())

    def test_importer_reads_storage_file_without_path_property(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "espak.xlsx"
            create_xlsx(path, rows=[["Hammer", 10, "", "", "", "SKU-1", "", ""]])
            parser_export = PathlessExport(self.shop, path)

            result = ExcelCatalogImporter().import_file(parser_export, column_map=EspakAdapter.column_map)

        self.assertEqual(result.products_created, 1)
        self.assertTrue(parser_export.import_success)
        self.assertEqual(ProductOffer.objects.get(shop=self.shop, external_id="SKU-1").original_name, "Hammer")

    @override_settings(MEDIA_ROOT=tempfile.mkdtemp())
    def test_fere_excel_import_creates_offer(self):
        fere_shop = Shop.objects.create(name="FERE", code="fere")
        fere_config = ParserConfig.objects.create(shop=fere_shop, name="FERE parser", code="fere")
        fere_run = ParserRun.objects.create(parser=fere_config, trigger=ParserRun.TRIGGER_COMMAND)

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "fere.xlsx"
            create_xlsx(
                path,
                headers=fere_parser.COLUMNS,
                rows=[['"EXPERT CUP" VÄRVINÕU SISU 5tk', 5.729, "", "", "4013307792031", "848351", "https://img.test/fere.jpg", "https://fere.ee/item"]],
            )
            parser_export = ParserExport(
                parser_run=fere_run,
                shop=fere_shop,
                original_filename=path.name,
                rows_count=1,
                file_size=path.stat().st_size,
            )
            with open(path, "rb") as handle:
                parser_export.file.save(path.name, File(handle), save=True)

            result = ExcelCatalogImporter().import_file(parser_export, column_map=FereAdapter.column_map)

        offer = ProductOffer.objects.get(shop=fere_shop, external_id="848351")
        self.assertEqual(result.products_created, 1)
        self.assertEqual(offer.original_name, '"EXPERT CUP" VÄRVINÕU SISU 5tk')
        self.assertEqual(str(offer.price), "5.73")
        self.assertEqual(offer.barcode, "4013307792031")

    def _create_export(self, path, rows_count):
        parser_export = ParserExport(
            parser_run=self.run,
            shop=self.shop,
            original_filename=path.name,
            rows_count=rows_count,
            file_size=path.stat().st_size,
        )
        with open(path, "rb") as handle:
            parser_export.file.save(path.name, File(handle), save=True)
        return parser_export


class FereAdapterTests(TestCase):
    def test_adapter_runs_standalone_with_output_path_and_log_callback(self):
        calls = {}

        async def fake_main(output_path=None, log_callback=None):
            calls["output_path"] = output_path
            calls["log_callback"] = log_callback
            log_callback("FERE progress")
            create_xlsx(
                output_path,
                headers=fere_parser.COLUMNS,
                rows=[["Hammer", 10, "", "", "4740000000001", "SKU-FERE", "https://img.test/fere.jpg", "https://fere.ee/item"]],
            )

        logs = []
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "fere.xlsx"
            with patch("parsers.adapters.fere.fere_parser.main", fake_main):
                result = asyncio.run(FereAdapter().run(output_path, log_callback=logs.append))

        self.assertTrue(result.success)
        self.assertEqual(result.products_count, 1)
        self.assertEqual(calls["output_path"], output_path)
        self.assertEqual(logs, ["FERE progress"])


class EhituseABCAdapterTests(TestCase):
    def test_standalone_wrapper_creates_excel(self):
        products = [
            {
                "sku": "A065162",
                "name": "6-KANTVÕTMETE KOMPLEKT 16 OSA KREATOR",
                "price": 25.9,
                "salePrice": 16.9,
                "image": "/images/product.jpeg",
                "url": "/ee/product",
            }
        ]
        logs = []

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "ehituseabc.xlsx"
            with patch("parsers.standalone.ehituseabc_parser.download_all_products", return_value=products):
                asyncio.run(ehituseabc_parser.main(output_path=output_path, log_callback=logs.append))

            workbook = load_workbook(output_path, read_only=True, data_only=True)
            try:
                worksheet = workbook[EhituseABCAdapter.worksheet_name]
                rows_count = max(worksheet.max_row - 1, 0)
            finally:
                workbook.close()

        self.assertEqual(rows_count, 1)
        self.assertTrue(logs)

    def test_adapter_runs_standalone_and_counts_products(self):
        calls = {}

        async def fake_main(output_path=None, log_callback=None):
            calls["output_path"] = output_path
            calls["log_callback"] = log_callback
            log_callback("EhituseABC progress")
            create_xlsx(
                output_path,
                headers=ehituseabc_parser.COLUMNS,
                sheet_name=EhituseABCAdapter.worksheet_name,
                rows=[
                    [
                        "Hammer",
                        10,
                        "",
                        "",
                        "4740000000001",
                        "ABC-1",
                        "https://img.test/abc.jpg",
                        "https://www.ehituseabc.ee/ee/item",
                    ]
                ],
            )

        logs = []
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "ehituseabc.xlsx"
            with patch("parsers.adapters.ehituseabc.ehituseabc_parser.main", fake_main):
                result = asyncio.run(EhituseABCAdapter().run(output_path, log_callback=logs.append))

        self.assertTrue(result.success)
        self.assertEqual(result.output_path, str(output_path))
        self.assertEqual(result.products_count, 1)
        self.assertEqual(calls["output_path"], output_path)
        self.assertEqual(logs, ["EhituseABC progress"])


class BauhausAdapterTests(TestCase):
    def setUp(self):
        self.shop = Shop.objects.create(name="BAUHAUS", code="bauhaus")
        self.config = ParserConfig.objects.create(
            shop=self.shop,
            name="BAUHAUS parser",
            code="bauhaus",
            run_order=6,
        )

    def test_standalone_wrapper_creates_excel_without_barcodes(self):
        products = [
            {
                "sku": "SKU-BAUHAUS-1",
                "name": "BAUHAUS hammer",
                "price": 14.95,
                "ordinary_price": 19.95,
                "image_url": "https://img.test/bauhaus.jpg",
                "product_url": "https://www.bauhaus.ee/item",
            }
        ]
        logs = []

        async def fake_request_text(session, url, semaphore):
            return "<html></html>"

        async def fake_discover_categories(session, semaphore, home):
            return [{"url": "https://www.bauhaus.ee/category"}]

        async def fake_collect_catalog(session, semaphore, categories):
            return products, {}

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "bauhaus.xlsx"
            with patch("parsers.standalone.bauhaus_parser.request_text", fake_request_text):
                with patch("parsers.standalone.bauhaus_parser.discover_top_level_categories", fake_discover_categories):
                    with patch("parsers.standalone.bauhaus_parser.collect_full_catalog", fake_collect_catalog):
                        with patch("parsers.standalone.bauhaus_parser.enrich_all_products_with_ean") as enrich_ean:
                            asyncio.run(bauhaus_parser.main(output_path=output_path, log_callback=logs.append))

            workbook = load_workbook(output_path, read_only=True, data_only=True)
            try:
                worksheet = workbook[BauhausAdapter.worksheet_name]
                rows_count = max(worksheet.max_row - 1, 0)
                barcode = worksheet["E2"].value
                external_id = worksheet["F2"].value
            finally:
                workbook.close()

        self.assertEqual(rows_count, 1)
        self.assertIn(barcode, (None, ""))
        self.assertEqual(external_id, "SKU-BAUHAUS-1")
        enrich_ean.assert_not_called()
        self.assertTrue(logs)

    def test_bauhaus_list_sku_exports_primary_identifier(self):
        products = [
            {
                "sku": ["PC692956", "1558895", "1558894"],
                "name": "BAUHAUS grouped product",
                "price": 2.9,
                "product_url": "https://www.bauhaus.ee/grouped-product",
            }
        ]

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "bauhaus.xlsx"
            with patch.object(bauhaus_parser, "OUTPUT_FILE", output_path):
                bauhaus_parser.export_final_excel(products)

            workbook = load_workbook(output_path, read_only=True, data_only=True)
            try:
                external_id = workbook[BauhausAdapter.worksheet_name]["F2"].value
            finally:
                workbook.close()

        self.assertEqual(external_id, "PC692956")

    def test_ean_error_does_not_prevent_excel_creation(self):
        products = [
            {
                "sku": "SKU-BAUHAUS-2",
                "name": "BAUHAUS drill",
                "price": 99.95,
                "product_url": "https://www.bauhaus.ee/drill",
            }
        ]

        async def fake_request_text(session, url, semaphore):
            return "<html></html>"

        async def fake_discover_categories(session, semaphore, home):
            return ["https://www.bauhaus.ee/category"]

        async def fake_collect_catalog(session, semaphore, categories):
            return products, {}

        with tempfile.TemporaryDirectory() as tmp_dir:
            with override_settings(
                MEDIA_ROOT=Path(tmp_dir) / "media",
                PARSER_EXPORT_WORK_DIR=Path(tmp_dir) / "work",
            ):
                with patch("parsers.standalone.bauhaus_parser.request_text", fake_request_text):
                    with patch("parsers.standalone.bauhaus_parser.discover_top_level_categories", fake_discover_categories):
                        with patch("parsers.standalone.bauhaus_parser.collect_full_catalog", fake_collect_catalog):
                            with patch("parsers.standalone.bauhaus_parser.enrich_all_products_with_ean") as enrich_ean:
                                with patch(
                                    "parsers.services.bauhaus_barcode_enricher.enrich_bauhaus_offer_barcodes",
                                    side_effect=RuntimeError("EAN endpoint unavailable"),
                                ) as incremental_enrichment:
                                    parser_run = run_excel_parser(self.config)

                parser_export = parser_run.export
                self.assertTrue(parser_export.file.storage.exists(parser_export.file.name))

        self.assertEqual(parser_run.status, ParserRun.STATUS_SUCCESS)
        self.assertTrue(parser_export.import_success)
        self.assertEqual(parser_run.excel_rows_count, 1)
        self.assertTrue(
            ProductOffer.objects.filter(
                shop=self.shop,
                external_id="SKU-BAUHAUS-2",
            ).exists()
        )
        created_offer = ProductOffer.objects.get(shop=self.shop, external_id="SKU-BAUHAUS-2")
        enrich_ean.assert_not_called()
        incremental_enrichment.assert_called_once()
        self.assertEqual(incremental_enrichment.call_args.args[0], [created_offer.pk])
        self.assertIn("WARNING: BAUHAUS barcode enrichment did not complete", parser_run.log)

    def test_bauhaus_excel_without_barcodes_passes_validation(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "bauhaus.xlsx"
            create_xlsx(
                output_path,
                headers=bauhaus_parser.COLUMNS,
                sheet_name=BauhausAdapter.worksheet_name,
                rows=[
                    [
                        "BAUHAUS hammer",
                        14.95,
                        "",
                        "",
                        "",
                        "SKU-BAUHAUS-1",
                        "https://img.test/bauhaus.jpg",
                        "https://www.bauhaus.ee/item",
                    ]
                ],
            )

            result = ExcelCatalogValidator().validate(
                output_path,
                column_map=BauhausAdapter.column_map,
                worksheet_name=BauhausAdapter.worksheet_name,
            )

        self.assertTrue(result.is_valid, result.error_message)
        self.assertEqual(result.rows_count, 1)

    def test_bauhaus_excel_import_creates_offer_without_barcode(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "bauhaus.xlsx"
            self._create_bauhaus_excel(output_path, name="BAUHAUS hammer")

            result = ExcelCatalogImporter().import_file(
                PathlessExport(self.shop, output_path),
                column_map=BauhausAdapter.column_map,
                worksheet_name=BauhausAdapter.worksheet_name,
            )

        offer = ProductOffer.objects.get(shop=self.shop, external_id="SKU-BAUHAUS-1")
        self.assertEqual(result.products_created, 1)
        self.assertEqual(result.created_offer_ids, [offer.pk])
        self.assertEqual(offer.original_name, "BAUHAUS hammer")
        self.assertEqual(offer.sku, "SKU-BAUHAUS-1")
        self.assertEqual(offer.barcode, "")

    def test_repeated_bauhaus_import_updates_offer_without_duplicates(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "bauhaus.xlsx"
            self._create_bauhaus_excel(output_path, name="BAUHAUS hammer")
            parser_export = PathlessExport(self.shop, output_path)
            first_result = ExcelCatalogImporter().import_file(
                parser_export,
                column_map=BauhausAdapter.column_map,
                worksheet_name=BauhausAdapter.worksheet_name,
            )
            offer = ProductOffer.objects.get(shop=self.shop, external_id="SKU-BAUHAUS-1")
            offer.barcode = "4006381333931"
            offer.product.barcode = "4006381333931"
            offer.product.save(update_fields=["barcode"])
            offer.save(update_fields=["barcode"])

            self._create_bauhaus_excel(output_path, name="BAUHAUS hammer updated", price=15.95)
            result = ExcelCatalogImporter().import_file(
                PathlessExport(self.shop, output_path),
                column_map=BauhausAdapter.column_map,
                worksheet_name=BauhausAdapter.worksheet_name,
            )

        offers = ProductOffer.objects.filter(shop=self.shop, external_id="SKU-BAUHAUS-1")
        self.assertEqual(offers.count(), 1)
        self.assertEqual(first_result.created_offer_ids, [offers.get().pk])
        self.assertEqual(result.created_offer_ids, [])
        self.assertEqual(result.products_created, 0)
        self.assertEqual(result.products_updated, 1)
        self.assertEqual(offers.get().original_name, "BAUHAUS hammer updated")
        self.assertEqual(offers.get().barcode, "4006381333931")
        self.assertEqual(offers.get().product.barcode, "4006381333931")

    def test_existing_bauhaus_offer_is_not_sent_to_incremental_enrichment(self):
        product = Product.objects.create(name="BAUHAUS drill")
        ProductOffer.objects.create(
            shop=self.shop,
            product=product,
            external_id="SKU-BAUHAUS-1",
            sku="SKU-BAUHAUS-1",
            original_name="BAUHAUS drill",
            price="10.00",
            product_url="https://www.bauhaus.ee/item",
        )
        products = [
            {
                "sku": "SKU-BAUHAUS-1",
                "name": "BAUHAUS drill updated",
                "price": 12.95,
                "product_url": "https://www.bauhaus.ee/item",
            }
        ]

        async def fake_request_text(session, url, semaphore):
            return "<html></html>"

        async def fake_discover_categories(session, semaphore, home):
            return ["https://www.bauhaus.ee/category"]

        async def fake_collect_catalog(session, semaphore, categories):
            return products, {}

        with tempfile.TemporaryDirectory() as tmp_dir:
            with override_settings(
                MEDIA_ROOT=Path(tmp_dir) / "media",
                PARSER_EXPORT_WORK_DIR=Path(tmp_dir) / "work",
            ):
                with patch("parsers.standalone.bauhaus_parser.request_text", fake_request_text):
                    with patch("parsers.standalone.bauhaus_parser.discover_top_level_categories", fake_discover_categories):
                        with patch("parsers.standalone.bauhaus_parser.collect_full_catalog", fake_collect_catalog):
                            with patch(
                                "parsers.services.bauhaus_barcode_enricher.enrich_bauhaus_offer_barcodes"
                            ) as incremental_enrichment:
                                parser_run = run_excel_parser(self.config)

        self.assertEqual(parser_run.status, ParserRun.STATUS_SUCCESS)
        incremental_enrichment.assert_not_called()
        offer = ProductOffer.objects.get(shop=self.shop, external_id="SKU-BAUHAUS-1")
        self.assertEqual(str(offer.price), "12.95")

    @staticmethod
    def _create_bauhaus_excel(path, *, name, price=14.95):
        create_xlsx(
            path,
            headers=bauhaus_parser.COLUMNS,
            sheet_name=BauhausAdapter.worksheet_name,
            rows=[
                [
                    name,
                    price,
                    "",
                    "",
                    "",
                    "SKU-BAUHAUS-1",
                    "https://img.test/bauhaus.jpg",
                    "https://www.bauhaus.ee/item",
                ]
            ],
        )

    def test_adapter_runs_standalone_and_counts_products(self):
        calls = {}

        async def fake_main(output_path=None, log_callback=None):
            calls["output_path"] = output_path
            calls["log_callback"] = log_callback
            log_callback("BAUHAUS progress")
            create_xlsx(
                output_path,
                headers=bauhaus_parser.COLUMNS,
                sheet_name=BauhausAdapter.worksheet_name,
                rows=[
                    [
                        "BAUHAUS hammer",
                        14.95,
                        12.95,
                        "",
                        "4740000000001",
                        "SKU-BAUHAUS-1",
                        "https://img.test/bauhaus.jpg",
                        "https://www.bauhaus.ee/item",
                    ]
                ],
            )

        logs = []
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "bauhaus.xlsx"
            with patch("parsers.adapters.bauhaus.bauhaus_parser.main", fake_main):
                result = asyncio.run(BauhausAdapter().run(output_path, log_callback=logs.append))

        self.assertTrue(result.success)
        self.assertEqual(result.output_path, str(output_path))
        self.assertEqual(result.products_count, 1)
        self.assertEqual(calls["output_path"], output_path)
        self.assertEqual(logs, ["BAUHAUS progress"])


class BauhofAdapterTests(TestCase):
    def test_standalone_wrapper_creates_excel(self):
        products = {
            "SKU-BH-1": {
                "Название товара": "Bauhof hammer",
                "Цена": 12.5,
                "Цена со скидкой": "",
                "Цена со скидкой 2": "",
                "Штрихкод": "4740000000001",
                "Код магазина": "SKU-BH-1",
                "Фото": "https://img.test/bauhof.jpg",
                "Ссылка": "https://www.bauhof.ee/item",
            }
        }
        logs = []

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "bauhof.xlsx"
            with patch("parsers.standalone.bauhof_parser.get_sku_map", return_value={"SKU-BH-1": "https://www.bauhof.ee/item"}):
                with patch("parsers.standalone.bauhof_parser.collect_products", return_value=products):
                    asyncio.run(bauhof_parser.main(output_path=output_path, log_callback=logs.append))

            workbook = load_workbook(output_path, read_only=True, data_only=True)
            try:
                worksheet = workbook[BauhofAdapter.worksheet_name]
                rows_count = max(worksheet.max_row - 1, 0)
            finally:
                workbook.close()

        self.assertEqual(rows_count, 1)
        self.assertTrue(logs)

    def test_adapter_runs_standalone_and_counts_products(self):
        calls = {}

        async def fake_main(output_path=None, log_callback=None):
            calls["output_path"] = output_path
            calls["log_callback"] = log_callback
            log_callback("Bauhof progress")
            create_xlsx(
                output_path,
                headers=bauhof_parser.COLUMNS,
                sheet_name=BauhofAdapter.worksheet_name,
                rows=[
                    [
                        "Bauhof hammer",
                        12.5,
                        "",
                        "",
                        "4740000000001",
                        "SKU-BH-1",
                        "https://img.test/bauhof.jpg",
                        "https://www.bauhof.ee/item",
                    ]
                ],
            )

        logs = []
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "bauhof.xlsx"
            with patch("parsers.adapters.bauhof.bauhof_parser.main", fake_main):
                result = asyncio.run(BauhofAdapter().run(output_path, log_callback=logs.append))

        self.assertTrue(result.success)
        self.assertEqual(result.output_path, str(output_path))
        self.assertEqual(result.products_count, 1)
        self.assertEqual(calls["output_path"], output_path)
        self.assertEqual(logs, ["Bauhof progress"])


class DepoAdapterTests(TestCase):
    def test_product_page_keeps_quantity_price_and_threshold(self):
        parser = depo_parser.DepoParser()
        parser.post_graphql = AsyncMock(
            return_value={
                "data": {
                    "products": {
                        "pageInfo": {"totalCount": 1},
                        "edges": [
                            {
                                "node": {
                                    "id": "DEPO-1",
                                    "name": "DEPO construction screw",
                                    "primaryBarcode": "4740000000001",
                                    "prices": {
                                        "yellow": {"priceWithVat": 10},
                                        "orange": {"priceWithVat": 8.5, "priceQuantity": 6},
                                    },
                                }
                            }
                        ],
                    }
                }
            }
        )

        total, rows = asyncio.run(parser.get_products_page(None, 8570, 0))

        self.assertEqual(total, 1)
        self.assertEqual(rows[0][depo_parser.COLUMNS[3]], 8.5)
        self.assertEqual(rows[0][depo_parser.COLUMNS[8]], 6)

    def test_first_category_pages_are_processed_by_the_worker_queue(self):
        async def scenario():
            parser = depo_parser.DepoParser()
            queue = asyncio.Queue()

            async def fake_page(session, category_id, start):
                total = 40 if category_id == 1 else 20
                return total, [
                    {
                        depo_parser.COLUMNS[0]: f"Product {category_id}-{start}",
                        depo_parser.COLUMNS[1]: 10,
                        depo_parser.COLUMNS[2]: "",
                        depo_parser.COLUMNS[3]: "",
                        depo_parser.COLUMNS[4]: "",
                        depo_parser.COLUMNS[5]: f"{category_id}-{start}",
                        depo_parser.COLUMNS[6]: "",
                        depo_parser.COLUMNS[7]: "",
                        depo_parser.COLUMNS[8]: "",
                    }
                ]

            parser.get_products_page = AsyncMock(side_effect=fake_page)
            await parser.prepare_queue(None, [1, 2], queue)
            workers = [asyncio.create_task(parser.worker(number, None, queue)) for number in (1, 2)]
            await asyncio.wait_for(queue.join(), timeout=1)
            for worker in workers:
                worker.cancel()
            await asyncio.gather(*workers, return_exceptions=True)
            return parser

        parser = asyncio.run(scenario())

        self.assertEqual(parser.categories_done, 2)
        self.assertEqual(parser.pages_done, 3)
        self.assertEqual(len(parser.products), 3)
        self.assertFalse(parser.errors)

    def test_standalone_wrapper_creates_excel(self):
        class FakeDepoParser(depo_parser.DepoParser):
            async def get_categories(self, session):
                return [1]

            async def prepare_queue(self, session, categories, queue):
                self.products["SKU-DEPO-1"] = {
                    "Название товара": "DEPO hammer",
                    "Цена": 11.25,
                    "Цена со скидкой": "",
                    "Цена со скидкой 2": "",
                    "Штрихкод": "4740000000001",
                    "Код магазина": "SKU-DEPO-1",
                    "Фото": "https://img.test/depo.jpg",
                    "Ссылка": "https://online.depo.ee/product/SKU-DEPO-1",
                }

            async def worker(self, number, session, queue):
                task = await queue.get()
                queue.task_done()
                if task is None:
                    return

        logs = []
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "depo.xlsx"
            with patch("parsers.standalone.depo_parser.DepoParser", FakeDepoParser):
                asyncio.run(depo_parser.main(output_path=output_path, log_callback=logs.append))

            workbook = load_workbook(output_path, read_only=True, data_only=True)
            try:
                worksheet = workbook[DepoAdapter.worksheet_name]
                rows_count = max(worksheet.max_row - 1, 0)
            finally:
                workbook.close()

        self.assertEqual(rows_count, 1)
        self.assertTrue(logs)

    def test_adapter_runs_standalone_and_counts_products(self):
        calls = {}

        async def fake_main(output_path=None, log_callback=None):
            calls["output_path"] = output_path
            calls["log_callback"] = log_callback
            log_callback("DEPO progress")
            create_xlsx(
                output_path,
                headers=depo_parser.COLUMNS,
                sheet_name=DepoAdapter.worksheet_name,
                rows=[
                    [
                        "DEPO hammer",
                        11.25,
                        "",
                        "",
                        "4740000000001",
                        "SKU-DEPO-1",
                        "https://img.test/depo.jpg",
                        "https://online.depo.ee/product/SKU-DEPO-1",
                    ]
                ],
            )

        logs = []
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "depo.xlsx"
            with patch("parsers.adapters.depo.depo_parser.main", fake_main):
                result = asyncio.run(DepoAdapter().run(output_path, log_callback=logs.append))

        self.assertTrue(result.success)
        self.assertEqual(result.output_path, str(output_path))
        self.assertEqual(result.products_count, 1)
        self.assertEqual(calls["output_path"], output_path)
        self.assertEqual(logs, ["DEPO progress"])


class AdapterRegistryTests(TestCase):
    existing_production_codes = {"espak", "depo", "bauhof", "ehituseabc", "fere", "bauhaus", "handymann"}
    expected_production_codes = existing_production_codes | set(PUBLIC_COMMERCE_STORES)

    def test_registry_contains_exactly_all_production_parsers(self):
        self.assertEqual(set(ADAPTERS), self.expected_production_codes)
        self.assertIs(ADAPTERS["bauhaus"], BauhausAdapter)
        self.assertIs(ADAPTERS["bauhof"], BauhofAdapter)
        self.assertIs(ADAPTERS["depo"], DepoAdapter)
        self.assertIs(ADAPTERS["ehituseabc"], EhituseABCAdapter)
        self.assertIs(ADAPTERS["espak"], EspakAdapter)
        self.assertIs(ADAPTERS["fere"], FereAdapter)
        self.assertIs(ADAPTERS["handymann"], HandymannAdapter)

    def test_get_adapter_class_returns_each_production_adapter(self):
        from parsers.adapters.registry import get_adapter_class

        for code in self.expected_production_codes:
            self.assertIs(get_adapter_class(code), ADAPTERS[code])

    def test_setup_parsers_and_registry_are_consistent(self):
        call_command("setup_parsers", verbosity=0)

        configs = ParserConfig.objects.filter(code__in=self.expected_production_codes).order_by("run_order")
        enabled_public_codes = {
            code
            for code, store in PUBLIC_COMMERCE_STORES.items()
            if store.enabled_by_default
        }

        self.assertEqual(set(configs.values_list("code", flat=True)), self.expected_production_codes)
        self.assertEqual(
            set(configs.filter(is_enabled=True).values_list("code", flat=True)),
            self.existing_production_codes | enabled_public_codes,
        )
        self.assertFalse(configs.get(code="horden").is_enabled)
        self.assertEqual(
            list(configs.values_list("run_order", flat=True)),
            list(range(1, len(self.expected_production_codes) + 1)),
        )


class PathlessFile:
    def __init__(self, path):
        self.path_to_open = path

    def open(self, mode="rb"):
        return open(self.path_to_open, mode)


class PathlessExport(SimpleNamespace):
    def __init__(self, shop, path):
        super().__init__(
            shop=shop,
            file=PathlessFile(path),
            import_success=False,
            validation_error="",
            imported_at=None,
        )

    def save(self, update_fields=None):
        return None


class BatchRunnerTests(TestCase):
    def setUp(self):
        self.shop_a = Shop.objects.create(name="A", code="a")
        self.shop_b = Shop.objects.create(name="B", code="b")
        self.config_b = ParserConfig.objects.create(shop=self.shop_b, name="B parser", code="b", run_order=2)
        self.config_a = ParserConfig.objects.create(shop=self.shop_a, name="A parser", code="a", run_order=1)

    def test_parsers_run_in_run_order_and_continue_after_failure(self):
        calls = []

        def fake_run_excel_parser(parser_config, trigger):
            calls.append(parser_config.code)
            status = ParserRun.STATUS_FAILED if parser_config.code == "a" else ParserRun.STATUS_SUCCESS
            return ParserRun.objects.create(parser=parser_config, trigger=trigger, status=status)

        with patch("parsers.services.batch_runner.ADAPTERS", {"a": object, "b": object}):
            with patch("parsers.services.batch_runner.run_excel_parser", fake_run_excel_parser):
                batch = run_all_parsers()

        self.assertEqual(calls, ["a", "b"])
        self.assertEqual(batch.status, ParserBatch.STATUS_PARTIAL)

    def test_running_batch_blocks_second_batch(self):
        ParserBatch.objects.create(status=ParserBatch.STATUS_RUNNING)

        with self.assertRaises(ParserBatchAlreadyRunning):
            run_all_parsers()

    def test_queue_job_is_processed_once(self):
        job = ParserQueueJob.objects.create(parser_config=self.config_a)

        def fake_run_excel_parser(parser_config, trigger):
            return ParserRun.objects.create(parser=parser_config, trigger=trigger, status=ParserRun.STATUS_SUCCESS)

        with patch("parsers.services.batch_runner.run_excel_parser", fake_run_excel_parser):
            processed = process_next_queue_job()
            second = process_next_queue_job()

        job.refresh_from_db()
        self.assertEqual(processed.pk, job.pk)
        self.assertIsNone(second)
        self.assertEqual(job.status, ParserQueueJob.STATUS_SUCCESS)
        self.assertEqual(job.attempts, 1)

    def test_cancel_requested_prevents_validation_stage(self):
        class FakeAdapter:
            column_map = EspakAdapter.column_map
            worksheet_name = None

        def fake_run_adapter(adapter, tmp_path, log):
            create_xlsx(tmp_path, rows=[["Hammer", 10, "", "", "", "SKU-1", "", ""]])
            ParserRun.objects.filter(parser=self.config_a, status=ParserRun.STATUS_RUNNING).update(cancel_requested=True)
            return ParserResult(success=True, output_path=str(tmp_path), products_count=1)

        with patch("parsers.services.batch_runner.get_adapter_class", return_value=FakeAdapter):
            with patch("parsers.services.batch_runner._run_adapter", fake_run_adapter):
                with patch.object(ExcelCatalogValidator, "validate") as validate:
                    run = run_excel_parser(self.config_a)

        validate.assert_not_called()
        self.assertEqual(run.status, ParserRun.STATUS_CANCELLED)
        self.assertIn("cancelled", run.error_message)

    def test_same_parser_config_cannot_create_second_running_run(self):
        ParserRun.objects.create(
            parser=self.config_a,
            status=ParserRun.STATUS_RUNNING,
            trigger=ParserRun.TRIGGER_COMMAND,
        )

        class FakeAdapter:
            column_map = EspakAdapter.column_map
            worksheet_name = None

        with patch("parsers.services.batch_runner.get_adapter_class", return_value=FakeAdapter):
            run = run_excel_parser(self.config_a)

        self.assertEqual(run.status, ParserRun.STATUS_FAILED)
        self.assertIn("already running", run.error_message)
        self.assertEqual(ParserRun.objects.filter(parser=self.config_a, status=ParserRun.STATUS_RUNNING).count(), 1)

    def test_different_parser_config_can_run_while_another_parser_is_running(self):
        ParserRun.objects.create(
            parser=self.config_a,
            status=ParserRun.STATUS_RUNNING,
            trigger=ParserRun.TRIGGER_COMMAND,
        )

        class FakeAdapter:
            column_map = EspakAdapter.column_map
            worksheet_name = None

        def fake_run_adapter(adapter, tmp_path, log):
            create_xlsx(tmp_path, rows=[["Hammer", 10, "", "", "", "SKU-B", "", ""]])
            return ParserResult(success=True, output_path=str(tmp_path), products_count=1)

        with tempfile.TemporaryDirectory() as tmp_dir:
            with override_settings(MEDIA_ROOT=Path(tmp_dir) / "media", PARSER_EXPORT_WORK_DIR=Path(tmp_dir) / "work"):
                with patch("parsers.services.batch_runner.get_adapter_class", return_value=FakeAdapter):
                    with patch("parsers.services.batch_runner._run_adapter", fake_run_adapter):
                        run = run_excel_parser(self.config_b)

        self.assertEqual(run.status, ParserRun.STATUS_SUCCESS)
        self.assertEqual(ParserRun.objects.filter(parser=self.config_b, status=ParserRun.STATUS_RUNNING).count(), 0)


class ParserRecoveryTests(TestCase):
    def setUp(self):
        self.shop = Shop.objects.create(name="ESPAK", code="espak")
        self.config = ParserConfig.objects.create(shop=self.shop, name="ESPAK parser", code="espak")

    @override_settings(PARSER_STALE_RUN_MINUTES=30, PARSER_STALE_JOB_MINUTES=30, PARSER_STALE_BATCH_MINUTES=30)
    def test_stale_running_parser_run_is_marked_failed(self):
        old_time = timezone.now() - timedelta(minutes=31)
        parser_run = ParserRun.objects.create(
            parser=self.config,
            status=ParserRun.STATUS_RUNNING,
            stage=ParserRun.STAGE_PARSING,
            trigger=ParserRun.TRIGGER_COMMAND,
            started_at=old_time,
            heartbeat_at=old_time,
        )

        result = recover_stale_parser_state()

        parser_run.refresh_from_db()
        self.assertEqual(result.runs, 1)
        self.assertEqual(parser_run.status, ParserRun.STATUS_FAILED)
        self.assertEqual(parser_run.stage, ParserRun.STAGE_COMPLETED)
        self.assertEqual(parser_run.error_message, STALE_RUN_MESSAGE)
        self.assertIsNotNone(parser_run.finished_at)

    @override_settings(PARSER_STALE_RUN_MINUTES=30)
    def test_fresh_running_parser_run_is_not_changed(self):
        parser_run = ParserRun.objects.create(
            parser=self.config,
            status=ParserRun.STATUS_RUNNING,
            stage=ParserRun.STAGE_PARSING,
            trigger=ParserRun.TRIGGER_COMMAND,
            started_at=timezone.now() - timedelta(minutes=10),
            heartbeat_at=timezone.now(),
        )

        result = recover_stale_parser_state()

        parser_run.refresh_from_db()
        self.assertEqual(result.runs, 0)
        self.assertEqual(parser_run.status, ParserRun.STATUS_RUNNING)

    @override_settings(PARSER_STALE_RUN_MINUTES=30)
    def test_parser_config_can_run_again_after_stale_recovery(self):
        old_time = timezone.now() - timedelta(minutes=31)
        ParserRun.objects.create(
            parser=self.config,
            status=ParserRun.STATUS_RUNNING,
            trigger=ParserRun.TRIGGER_COMMAND,
            started_at=old_time,
            heartbeat_at=old_time,
        )

        recover_stale_parser_state()

        class FakeAdapter:
            column_map = EspakAdapter.column_map
            worksheet_name = None

        def fake_run_adapter(adapter, tmp_path, log):
            create_xlsx(tmp_path, rows=[["Hammer", 10, "", "", "", "SKU-RECOVERED", "", ""]])
            return ParserResult(success=True, output_path=str(tmp_path), products_count=1)

        with tempfile.TemporaryDirectory() as tmp_dir:
            with override_settings(MEDIA_ROOT=Path(tmp_dir) / "media", PARSER_EXPORT_WORK_DIR=Path(tmp_dir) / "work"):
                with patch("parsers.services.batch_runner.get_adapter_class", return_value=FakeAdapter):
                    with patch("parsers.services.batch_runner._run_adapter", fake_run_adapter):
                        parser_run = run_excel_parser(self.config)

        self.assertEqual(parser_run.status, ParserRun.STATUS_SUCCESS)

    @override_settings(PARSER_STALE_JOB_MINUTES=30)
    def test_stale_queue_job_is_marked_failed(self):
        old_time = timezone.now() - timedelta(minutes=31)
        job = ParserQueueJob.objects.create(
            parser_config=self.config,
            status=ParserQueueJob.STATUS_RUNNING,
            started_at=old_time,
            heartbeat_at=old_time,
        )

        result = recover_stale_parser_state()

        job.refresh_from_db()
        self.assertEqual(result.jobs, 1)
        self.assertEqual(job.status, ParserQueueJob.STATUS_FAILED)
        self.assertEqual(job.error_message, STALE_JOB_MESSAGE)
        self.assertIsNotNone(job.finished_at)

    @override_settings(PARSER_STALE_BATCH_MINUTES=30)
    def test_stale_batch_is_marked_failed(self):
        old_time = timezone.now() - timedelta(minutes=31)
        batch = ParserBatch.objects.create(
            status=ParserBatch.STATUS_RUNNING,
            current_parser=self.config,
            started_at=old_time,
            heartbeat_at=old_time,
            log="started",
        )

        result = recover_stale_parser_state()

        batch.refresh_from_db()
        self.assertEqual(result.batches, 1)
        self.assertEqual(batch.status, ParserBatch.STATUS_FAILED)
        self.assertIsNone(batch.current_parser)
        self.assertIn(STALE_BATCH_MESSAGE, batch.log)
        self.assertIsNotNone(batch.finished_at)

    @override_settings(PARSER_STALE_RUN_MINUTES=30, PARSER_STALE_JOB_MINUTES=30, PARSER_STALE_BATCH_MINUTES=30)
    def test_recover_stale_parser_jobs_command_outputs_statistics(self):
        old_time = timezone.now() - timedelta(minutes=31)
        ParserRun.objects.create(
            parser=self.config,
            status=ParserRun.STATUS_RUNNING,
            trigger=ParserRun.TRIGGER_COMMAND,
            started_at=old_time,
            heartbeat_at=old_time,
        )
        ParserQueueJob.objects.create(
            parser_config=self.config,
            status=ParserQueueJob.STATUS_RUNNING,
            started_at=old_time,
            heartbeat_at=old_time,
        )
        ParserBatch.objects.create(status=ParserBatch.STATUS_RUNNING, started_at=old_time, heartbeat_at=old_time)
        output = StringIO()

        call_command("recover_stale_parser_jobs", stdout=output)

        self.assertIn("runs=1", output.getvalue())
        self.assertIn("jobs=1", output.getvalue())
        self.assertIn("batches=1", output.getvalue())

    @override_settings(PARSER_STALE_RUN_MINUTES=30)
    def test_recovery_does_not_overwrite_run_completed_during_recovery(self):
        old_time = timezone.now() - timedelta(minutes=31)
        parser_run = ParserRun.objects.create(
            parser=self.config,
            status=ParserRun.STATUS_RUNNING,
            trigger=ParserRun.TRIGGER_COMMAND,
            started_at=old_time,
            heartbeat_at=old_time,
        )

        def complete_before_update(obj, cutoff):
            ParserRun.objects.filter(pk=parser_run.pk).update(status=ParserRun.STATUS_SUCCESS)
            return True

        with patch("parsers.services.recovery._is_stale", complete_before_update):
            result = recover_stale_parser_state()

        parser_run.refresh_from_db()
        self.assertEqual(result.runs, 0)
        self.assertEqual(parser_run.status, ParserRun.STATUS_SUCCESS)


class HeartbeatTickerTests(TransactionTestCase):
    def setUp(self):
        self.shop = Shop.objects.create(name="ESPAK", code="espak")
        self.config = ParserConfig.objects.create(shop=self.shop, name="ESPAK parser", code="espak")

    @override_settings(PARSER_STALE_JOB_MINUTES=1)
    def test_live_queue_job_with_heartbeat_is_not_recovered(self):
        old_time = timezone.now() - timedelta(minutes=2)
        job = ParserQueueJob.objects.create(
            parser_config=self.config,
            status=ParserQueueJob.STATUS_RUNNING,
            started_at=old_time,
            heartbeat_at=old_time,
        )

        with HeartbeatTicker([(ParserQueueJob, job.pk, ParserQueueJob.STATUS_RUNNING)], interval_seconds=0.05):
            time.sleep(0.15)
            result = recover_stale_parser_state()

        job.refresh_from_db()
        self.assertEqual(result.jobs, 0)
        self.assertEqual(job.status, ParserQueueJob.STATUS_RUNNING)
        self.assertGreater(job.heartbeat_at, old_time)

    @override_settings(PARSER_STALE_BATCH_MINUTES=1)
    def test_live_batch_with_heartbeat_is_not_recovered(self):
        old_time = timezone.now() - timedelta(minutes=2)
        batch = ParserBatch.objects.create(
            status=ParserBatch.STATUS_RUNNING,
            started_at=old_time,
            heartbeat_at=old_time,
        )

        with HeartbeatTicker([(ParserBatch, batch.pk, ParserBatch.STATUS_RUNNING)], interval_seconds=0.05):
            time.sleep(0.15)
            result = recover_stale_parser_state()

        batch.refresh_from_db()
        self.assertEqual(result.batches, 0)
        self.assertEqual(batch.status, ParserBatch.STATUS_RUNNING)
        self.assertGreater(batch.heartbeat_at, old_time)

    def test_parser_run_without_logs_receives_heartbeat(self):
        old_time = timezone.now() - timedelta(minutes=2)
        parser_run = ParserRun.objects.create(
            parser=self.config,
            status=ParserRun.STATUS_RUNNING,
            trigger=ParserRun.TRIGGER_COMMAND,
            started_at=old_time,
            heartbeat_at=old_time,
        )

        with HeartbeatTicker([(ParserRun, parser_run.pk, ParserRun.STATUS_RUNNING)], interval_seconds=0.05):
            time.sleep(0.15)

        parser_run.refresh_from_db()
        self.assertGreater(parser_run.heartbeat_at, old_time)

    def test_heartbeat_database_error_does_not_escape(self):
        parser_run = ParserRun.objects.create(
            parser=self.config,
            status=ParserRun.STATUS_RUNNING,
            trigger=ParserRun.TRIGGER_COMMAND,
            heartbeat_at=timezone.now(),
        )
        ticker = HeartbeatTicker([(ParserRun, parser_run.pk, ParserRun.STATUS_RUNNING)], interval_seconds=0.05)

        with patch("django.db.models.query.QuerySet.update", side_effect=OperationalError("database is locked")):
            ticker.beat()

    def test_ticker_stops_after_context_exit(self):
        old_time = timezone.now() - timedelta(minutes=2)
        job = ParserQueueJob.objects.create(
            parser_config=self.config,
            status=ParserQueueJob.STATUS_RUNNING,
            started_at=old_time,
            heartbeat_at=old_time,
        )

        with HeartbeatTicker([(ParserQueueJob, job.pk, ParserQueueJob.STATUS_RUNNING)], interval_seconds=0.05):
            time.sleep(0.12)

        job.refresh_from_db()
        stopped_at = job.heartbeat_at
        time.sleep(0.15)
        job.refresh_from_db()
        self.assertEqual(job.heartbeat_at, stopped_at)


class BatchConcurrencyTests(TransactionTestCase):
    def setUp(self):
        ParserBatchLock.objects.get_or_create(name="nightly_parser_batch")

    def test_two_parallel_start_batch_calls_do_not_create_two_running_batches(self):
        barrier = threading.Barrier(2)

        def call_start_batch():
            barrier.wait()
            try:
                return start_batch(ParserRun.TRIGGER_COMMAND).status
            except ParserBatchAlreadyRunning:
                return "blocked"

        with ThreadPoolExecutor(max_workers=2) as executor:
            results = list(executor.map(lambda _: call_start_batch(), range(2)))

        self.assertLessEqual(results.count(ParserBatch.STATUS_RUNNING), 1)
        self.assertGreaterEqual(results.count("blocked"), 1)
        self.assertLessEqual(ParserBatch.objects.filter(status=ParserBatch.STATUS_RUNNING).count(), 1)

    def test_missing_batch_lock_raises_configuration_error(self):
        ParserBatchLock.objects.filter(name="nightly_parser_batch").delete()

        with self.assertRaises(ParserBatchLockMissing):
            start_batch(ParserRun.TRIGGER_COMMAND)


class ParserExportAdminTests(TestCase):
    @override_settings(MEDIA_ROOT=tempfile.mkdtemp())
    def test_staff_with_permission_can_download_export(self):
        shop = Shop.objects.create(name="ESPAK", code="espak")
        config = ParserConfig.objects.create(shop=shop, name="ESPAK parser", code="espak")
        run = ParserRun.objects.create(parser=config, trigger=ParserRun.TRIGGER_COMMAND)
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "espak.xlsx"
            create_xlsx(path, rows=[["Hammer", 10, "", "", "", "SKU-1", "", ""]])
            parser_export = ParserExport(parser_run=run, shop=shop, original_filename=path.name, rows_count=1)
            with open(path, "rb") as handle:
                parser_export.file.save(path.name, File(handle), save=True)

        user = get_user_model().objects.create_superuser("admin", "admin@example.com", "password")
        client = Client()
        client.force_login(user)
        response = client.get(reverse("admin:parsers_parserexport_download", args=[parser_export.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment", response["Content-Disposition"])

    @override_settings(MEDIA_ROOT=tempfile.mkdtemp())
    def test_anonymous_user_cannot_download_export(self):
        shop = Shop.objects.create(name="ESPAK", code="espak")
        config = ParserConfig.objects.create(shop=shop, name="ESPAK parser", code="espak")
        run = ParserRun.objects.create(parser=config, trigger=ParserRun.TRIGGER_COMMAND)
        parser_export = ParserExport.objects.create(
            parser_run=run,
            shop=shop,
            original_filename="espak.xlsx",
            rows_count=1,
        )

        response = Client().get(reverse("admin:parsers_parserexport_download", args=[parser_export.pk]))

        self.assertEqual(response.status_code, 302)
