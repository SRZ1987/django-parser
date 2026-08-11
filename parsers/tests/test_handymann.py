import asyncio
import tempfile
from pathlib import Path
from unittest.mock import patch

from django.core.files import File
from django.test import SimpleTestCase, TestCase
from django.utils import timezone
from openpyxl import load_workbook

from catalog.models import ProductOffer, Shop
from parsers.adapters.handymann import HandymannAdapter
from parsers.models import ParserConfig, ParserExport, ParserRun
from parsers.services.excel_importer import ExcelCatalogImporter
from parsers.services.excel_validation import ExcelCatalogValidator
from parsers.standalone import handymann_parser


def api_product(
    product_id=101,
    *,
    sku="HM-101",
    name="Handymann hammer",
    price="1299",
    regular_price="1599",
    sale_price="1299",
    in_stock=True,
):
    return {
        "id": product_id,
        "name": name,
        "sku": sku,
        "permalink": f"https://handymann.ee/toode/{product_id}/",
        "is_in_stock": in_stock,
        "prices": {
            "price": price,
            "regular_price": regular_price,
            "sale_price": sale_price,
            "currency_minor_unit": 2,
        },
        "images": [{"src": f"https://img.test/{product_id}.jpg"}],
        "categories": [{"id": 9, "name": "Garden trimmers"}],
        "description": "Cutting width 25cm",
        "brands": [{"name": "Trolla"}],
        "attributes": [{"name": "Model", "terms": [{"name": "HM-350"}]}],
        "extensions": {},
    }


class HandymannParserUnitTests(SimpleTestCase):
    def test_product_normalization_preserves_sku_and_sale_price(self):
        product = api_product(name="Haamer &amp; naelad")

        row = handymann_parser.normalize_product(product)

        self.assertEqual(row[0], "Haamer & naelad")
        self.assertEqual(row[1], 15.99)
        self.assertEqual(row[2], 12.99)
        self.assertEqual(row[4], "")
        self.assertEqual(row[5], "HM-101")
        self.assertEqual(row[9], "Garden trimmers")
        self.assertEqual(row[11], "Cutting width 25cm")
        self.assertEqual(row[12], "Trolla")
        self.assertEqual(row[13], "HM-350")

    def test_product_without_sku_uses_stable_woocommerce_id(self):
        row = handymann_parser.normalize_product(api_product(product_id=777, sku=""))

        self.assertEqual(row[5], "wc-777")

    def test_out_of_stock_product_is_not_exported(self):
        self.assertIsNone(handymann_parser.normalize_product(api_product(in_stock=False)))

    def test_fetch_catalog_downloads_every_page(self):
        calls = []

        async def fake_request_page(_session, page, _log_callback=None):
            calls.append(page)
            headers = {"X-WP-Total": "2", "X-WP-TotalPages": "2"}
            return [api_product(product_id=page, sku=f"HM-{page}")], headers

        with patch("parsers.standalone.handymann_parser.request_page", fake_request_page):
            products = asyncio.run(handymann_parser.fetch_catalog())

        self.assertEqual({product["id"] for product in products}, {1, 2})
        self.assertEqual(sorted(calls), [1, 2])

    def test_incomplete_catalog_fails_before_excel_is_created(self):
        async def fake_request_page(_session, page, _log_callback=None):
            headers = {"X-WP-Total": "3", "X-WP-TotalPages": "2"}
            return [api_product(product_id=page, sku=f"HM-{page}")], headers

        with patch("parsers.standalone.handymann_parser.request_page", fake_request_page):
            with self.assertRaisesRegex(RuntimeError, "catalog is incomplete"):
                asyncio.run(handymann_parser.fetch_catalog())


class HandymannAdapterTests(SimpleTestCase):
    def test_standalone_creates_excel(self):
        products = [api_product(), api_product(product_id=102, sku="HM-102", in_stock=False)]

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "handymann.xlsx"
            with patch("parsers.standalone.handymann_parser.fetch_catalog", return_value=products):
                asyncio.run(handymann_parser.main(output_path=output_path))

            workbook = load_workbook(output_path, read_only=True, data_only=True)
            try:
                worksheet = workbook[HandymannAdapter.worksheet_name]
                self.assertEqual(worksheet.max_row - 1, 1)
                self.assertEqual(worksheet["F2"].value, "HM-101")
            finally:
                workbook.close()

    def test_adapter_runs_standalone_and_counts_products(self):
        calls = {}

        async def fake_main(output_path=None, log_callback=None):
            calls["output_path"] = output_path
            log_callback("Handymann progress")
            handymann_parser.save_excel(
                [handymann_parser.normalize_product(api_product())],
                Path(output_path),
            )

        logs = []
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "handymann.xlsx"
            with patch("parsers.adapters.handymann.handymann_parser.main", fake_main):
                result = asyncio.run(HandymannAdapter().run(output_path, log_callback=logs.append))

        self.assertTrue(result.success)
        self.assertEqual(result.output_path, str(output_path))
        self.assertEqual(result.products_count, 1)
        self.assertEqual(calls["output_path"], output_path)
        self.assertEqual(logs, ["Handymann progress"])


class HandymannExcelPipelineTests(TestCase):
    def setUp(self):
        self.shop = Shop.objects.create(name="Handymann", code="handymann", website_url=handymann_parser.BASE_URL)
        self.config = ParserConfig.objects.create(
            shop=self.shop,
            name="Handymann parser",
            code="handymann",
            run_order=7,
        )
        self.parser_run = ParserRun.objects.create(
            parser=self.config,
            status=ParserRun.STATUS_RUNNING,
            started_at=timezone.now(),
        )

    def test_excel_validates_and_imports_product_offer(self):
        row = handymann_parser.normalize_product(api_product())
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "handymann.xlsx"
            handymann_parser.save_excel([row], path)

            validation = ExcelCatalogValidator().validate(
                path,
                column_map=HandymannAdapter.column_map,
                worksheet_name=HandymannAdapter.worksheet_name,
            )
            self.assertTrue(validation.is_valid, validation.error_message)

            parser_export = ParserExport(
                parser_run=self.parser_run,
                shop=self.shop,
                original_filename=path.name,
                rows_count=validation.rows_count,
                file_size=path.stat().st_size,
            )
            with path.open("rb") as handle:
                parser_export.file.save(path.name, File(handle), save=True)

            result = ExcelCatalogImporter().import_file(
                parser_export,
                column_map=HandymannAdapter.column_map,
                worksheet_name=HandymannAdapter.worksheet_name,
            )

        offer = ProductOffer.objects.get(shop=self.shop, external_id="HM-101")
        self.assertEqual(result.products_created, 1)
        self.assertEqual(offer.sku, "HM-101")
        self.assertEqual(str(offer.price), "15.99")
        self.assertEqual(str(offer.sale_price), "12.99")
        self.assertEqual(offer.category.name, "Garden trimmers")
        self.assertEqual(offer.description, "Cutting width 25cm")
        self.assertEqual(offer.product.brand, "Trolla")
        self.assertEqual(offer.product.model, "HM-350")
